from helpers import SqlConnect
from models.musteriler_model import *
from openpyxl import *
import shutil

class MusteriIslem:


    def __init__(self):

        self.data = SqlConnect().data


    def getMusteriListesi(self):

        result = self.data.getList(
            """
            select
            m.ID,
            m.FirmaAdi,
            m.Unvan,
            m.Adres,
            m.Marketing,
            u.UlkeAdi,
            u.Png_Flags,
            k.KullaniciAdi as Temsilci,
            m.Devir,
            m.Ozel,
            m.Telefon,
            m.Sira,
			m.Notlar,
            m.SonKullanici,
			(select KullaniciAdi from KullaniciTB ku where ku.ID = m.Satisci) as Satisci
            from
            MusterilerTB m,YeniTeklif_UlkeTB u,KullaniciTB k
            where u.Id=m.UlkeId and k.ID=m.MusteriTemsilciId 
            order by m.ID
            """
        )

        liste = list()
        sira = 1
        for item in result:

            model = MusteriListeModel()
            model.id = item.ID
            model.musteriadi = item.FirmaAdi
            model.unvan = item.Unvan
            model.adres = item.Adres
            model.marketing = item.Marketing
            model.ulkeadi = item.UlkeAdi
            model.logo = item.Png_Flags
            model.temsilci = item.Temsilci
            model.devir = item.Devir
            model.ozel = item.Ozel
            model.telefon = item.Telefon
            model.sira = item.Sira
            model.satisci = item.Satisci
            model.sonkullanici = item.SonKullanici
            model.musteri_sira = sira
            sira += 1
            liste.append(model)

        schema = MusteriListeSchema(many=True)

        return schema.dump(liste)

    def getMusteriListesiYil(self,year):
        result = self.data.getStoreList(
            """
            select
            m.ID,
            m.FirmaAdi,
            m.Unvan,
            m.Adres,
            m.Marketing,
            u.UlkeAdi,
            u.Png_Flags,
            k.KullaniciAdi as Temsilci,
            m.Devir,
            m.Ozel,
            m.Telefon,
            m.Sira,
			m.Notlar,
            m.SonKullanici,
			(select KullaniciAdi from KullaniciTB ku where ku.ID = m.Satisci) as Satisci
            from
            MusterilerTB m,YeniTeklif_UlkeTB u,KullaniciTB k
            where u.Id=m.UlkeId and k.ID=m.MusteriTemsilciId and YEAR(m.KayitTarihi) = ?
            order by m.ID
            """,(year)
        )

        liste = list()
        sira = 1
        for item in result:

            model = MusteriListeModel()
            model.id = item.ID
            model.musteriadi = item.FirmaAdi
            model.unvan = item.Unvan
            model.adres = item.Adres
            model.marketing = item.Marketing
            model.ulkeadi = item.UlkeAdi
            model.logo = item.Png_Flags
            model.temsilci = item.Temsilci
            model.devir = item.Devir
            model.ozel = item.Ozel
            model.telefon = item.Telefon
            model.sira = item.Sira
            model.satisci = item.Satisci
            model.sonkullanici = item.SonKullanici
            model.musteri_sira = sira
            sira += 1
            liste.append(model)

        schema = MusteriListeSchema(many=True)

        return schema.dump(liste)



    def getMusteriYilListesi(self):
        result = self.data.getList("""
                                    select YEAR(SiparisTarihi) as Yil from SiparislerTB group by YEAR(SiparisTarihi) order by YEAR(SiparisTarihi) desc
                                   """)
        liste = list()
        for item in result:
            model = MusteriSipYilListModel()
            model.yil = item.Yil
            liste.append(model)
        schema = MusteriSipYilListSchema(many=True)
        return schema.dump(liste)

        
    
    
    
    
    
    def excelCiktiAl(self,data_list):


        source_path = 'resource_api/musteriler/sablonlar/musteriDetayListesi.xlsx'
        target_path = 'resource_api/musteriler/dosyalar/musteriDetayListesi.xlsx'
        
        shutil.copy2(source_path, target_path)

        kitap = load_workbook(target_path)
        sayfa = kitap.get_sheet_by_name('Sayfa1')

        satir = 2

        musteri_listesi = data_list

        for item in musteri_listesi:

            sayfa.cell(satir,column=1,value=item['id'])
            sayfa.cell(satir,column=2,value=item['musteriadi'])
            sayfa.cell(satir,column=3,value=item['unvan'])
            sayfa.cell(satir,column=4,value=item['adres'])
            sayfa.cell(satir,column=5,value=item['marketing'])
            sayfa.cell(satir,column=6,value=item['ulkeadi'])
            sayfa.cell(satir,column=7,value=item['telefon'])
            sayfa.cell(satir,column=8,value=item['temsilci'])
            sayfa.cell(satir,column=9,value=item['devir'])
            sayfa.cell(satir,column=10,value=item['ozel'])

            satir += 1
        kitap.save(target_path)
        kitap.close()

        return target_path

    def setCustomerFollowing(self,customer,follow):
        try:
            if follow=='true':
                follow = True
            else:
                follow = False
            self.data.update_insert("""
                                        update MusterilerTB SET Takip=? where FirmaAdi=?
                                    
                                    
                                    """,(follow,customer))
            return True
        except Exception as e:
            print('setCustomerFollowing',str(e))
            return False

    def setSurfaceCustomers(self,data):
        try:
            surface_id = self.__setCountryId(data['surface'],data['user_id'])
            self.data.update_insert("""
                                        insert into SurfaceCustomersTB(FirstName,Adress,City,Email,Phone,SurfaceId,UserId) VALUES(?,?,?,?,?,?,?)
                                    
                                    """,(data['name'],data['adress'],data['city'],data['email'],data['phone'],surface_id,data['user_id']))
            return True
        except Exception as e:
            print("setSurfaceCustomers hata",str(e))
            return False

    def setSurfaceCustomersUpdate(self,data):
        try:
            surface_id = self.__setCountryId(data['surface'],data['user_id'])
            self.data.update_insert("""
                                        update SurfaceCustomersTB SET FirstName=?,Adress=?,City=?,Email=?,Phone=?,SurfaceId=? Where ID=?
                                    
                                    """,(data['name'],data['adress'],data['city'],data['email'],data['phone'],surface_id,data['id']))
            
            
            surfaceID = self.data.getStoreList("""
                                                    select ID from CustomersSurfaceTB where Surface=?

                                               """,(data['oldSurface']))[0].ID
            result = self.data.getStoreList("""
                                                select * from SurfaceCustomersTB where SurfaceId=?
                                            
                                            """,(surfaceID))
            if len(result)==0:
                self.data.update_insert("""
                                        delete CustomersSurfaceTB where ID=?
                                       
                                       """,(surfaceID))
            return True
        except Exception as e:
            print("setSurfaceCustomersUpdate hata",str(e))
            return False
            
    def setSurfaceCustomersDelete(self,id):
        try:
            surface =  self.data.getStoreList("""
                                        select SurfaceId from SurfaceCustomersTB where ID=?

                                   """,(id))[0].SurfaceId
            
                        
            
            self.data.update_insert("""
                                        delete SurfaceCustomersTB where ID=?
                                    
                                    """,(id))
            
            
            result = self.data.getStoreList("""
                                                select * from SurfaceCustomersTB where SurfaceId=?
                                            
                                            """,(surface))
            if len(result)==0:
                self.data.update_insert("""
                                        delete CustomersSurfaceTB where ID=?
                                       
                                       """,(surface))
            
            
            
            return True
        except Exception as e:
            print("setSurfaceCustomersDelete hata",str(e))
            return False
        
        
        
        
    def getCustomerSurfaceList(self,user_id):
        try:
            result = self.data.getStoreList("""
                                            select 

                                                sc.*,
                                                cs.Surface

                                            from SurfaceCustomersTB sc
                                            inner join CustomersSurfaceTB cs on sc.SurfaceId = cs.ID
											where sc.UserId=?
                                       
                                       """,(user_id))
            liste = list()
            for item in result:
                model = CustomersSurfaceListModel()
                model.id = item.ID
                model.surface = item.Surface
                model.firstName = item.FirstName
                model.adress = item.Adress
                model.city = item.City
                model.email = item.Email
                model.phone = item.Phone
                model.surfaceId = item.SurfaceId
                model.user_id = item.UserId
                
                liste.append(model)
            schema = CustomersSurfaceListSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print("getCustomerSurfaceList hata",str(e))
            return False
    
    def getSurfaceList(self,user_id):
        try:
            result = self.data.getStoreList("""
                                            select * from CustomersSurfaceTB where UserId=?
                                       
                                       
                                       """,(user_id))
            liste = list()
            for item in result:
                model = CustomersSurfaceListModel()
                model.id = item.ID
                model.surface = item.Surface
                liste.append(model)
            schema = CustomersSurfaceListSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print('getSurface hata',str(e))
            return False
    
    
    def getTekliflerMusteriListesi(self):
        try:
            result = self.data.getList("""
                                    select 

                                    ym.*,
                                    (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = ym.UlkeId) as UlkeAdi

                                from YeniTeklif_MusterilerTB ym
                              """)
            liste = list()
            for item in result:
                model = CustomersSurfaceListModel()
                model.id = item.Id
                musteri_adi = item.MusteriAdi
                musteri = musteri_adi.split('-')
                if(len(musteri)>0):
                    model.firstName = musteri[0].strip()
                else:
                    model.firstName = item.MusteriAdi.strip()

                
                model.adress = self.__noneControl(item.Adress)
                model.city = self.__noneControl(item.UlkeAdi)
                model.email = self.__noneControl(item.Mail)
                model.phone = self.__noneControl(item.Phone)
                liste.append(model)
            schema = CustomersSurfaceListSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print('getTekliflerMusteriListesi hata',str(e))
            return False

    def __noneControl(self,value):
        if(value is not None):
            return value
        else:
            return ""
    
    def __setCountryId(self,surface,user_id):
        try:
            surface = surface.capitalize()
            result = self.data.getStoreList("""
                                    select * from CustomersSurfaceTB where Surface =?
                                   
                                   """,(surface))
            if(len(result)>0):
                return result[0].ID
            else:
                self.data.update_insert("""
                                            insert into CustomersSurfaceTB(Surface,UserId) VALUES(?,?)
                                        
                                        """,(surface,user_id))
                
                result = self.data.getStoreList("""
                                        select * from CustomersSurfaceTB where Surface =?
                                    
                                    """,(surface))
                
                return result[0].ID
            
        except Exception as e:
            print('__setCountryId hata ',str(e))
            return False

class MusteriSiparisIslem:
    def __init__(self):

        self.data = SqlConnect().data

    def getMusteriSiparisListesi(self):
        result = self.data.getList(
            """
             Select 
                u.SiparisNo , 
                u.SatisFiyati , 
                s.SiparisTarihi,
                (select m.FirmaAdi from  MusterilerTB m where  m.ID=s.MusteriID) as musteri,
                (select r.UrunAdi from  UrunlerTB r where r.ID= k.UrunID ) as urunadi,
                (select y.YuzeyIslemAdi from  YuzeyKenarTB y where y.ID= k.YuzeyID ) as yuzeyadi,
                (select o.En from  OlculerTB o where o.ID= k.OlcuID ) as en,
                (select o.Boy from  OlculerTB o where o.ID= k.OlcuID ) as boy,
                (select o.Kenar from  OlculerTB o where o.ID= k.OlcuID ) as kenar,
				(select ub.BirimAdi from UrunBirimTB ub WHERE ub.ID = u.UrunBirimID) as urunbirim
				

                from SiparisUrunTB u , UrunKartTB k , SiparislerTB s 

                where u.UrunKartID = k.ID
                and s.SiparisNo  = u.SiparisNo
                and Year(s.SiparisTarihi) in (2023,2022,2021)
                and s.SiparisDurumID=3
                and u.SatisFiyati != 0
                order by s.SiparisTarihi desc


            """
        )

        liste = list()
        for item in result:

            model = MusteriSiparisListeModel()
            model.firmaadi = item.musteri
            model.urunadi = item.urunadi
            model.satisFiyati = item.SatisFiyati
            model.siparisno = item.SiparisNo
            model.yuzeyadi = item.yuzeyadi
            model.en = item.en
            model.boy = item.boy
            model.kenar = item.kenar
            model.year = item.SiparisTarihi
            model.urunbirim = item.urunbirim
            

            liste.append(model)

        schema = MusteriSiparisListeSchema(many=True)

        return schema.dump(liste)


class MusteriSiparisAyrintiCardIslem:
    def __init__(self):
        self.data = SqlConnect().data
        
    def getMusteriSiparisAyrintiCard(self):
        sumResult = self.data.getList("""
                                        select 
                                            m.ID,
                                            m.FirmaAdi,
                                            sum(su.SatisToplam) as SumOrder,
                                            m.Marketing

                                        from MusterilerTB m
                                            inner join SiparisUrunTB su on su.musteriID=m.ID

                                            group by m.ID,m.FirmaAdi,m.Marketing
                                      """)
        liste = list()
        for item in sumResult:
            model = MusteriSipAyrintiCardModel()
            model.firmaAdi = item.FirmaAdi
            model.sumOrder = item.SumOrder
            model.topOrder = self.getMusteriSipTop(item.ID)
            model.marketing = item.Marketing
            
            liste.append(model)
        
        schema =MusteriSipAyrintiCardSchema(many=True)
        return schema.dump(liste)
        
        
        
        
    
    def getMusteriSipTop(self,id):
        
        topResult = self.data.getStoreList("""
                                            select 

                                                TOP 1
                                                
                                                su.SiparisNo

                                            from MusterilerTB m
                                                inner join SiparislerTB su on su.musteriID=m.ID

                                                where m.ID=?

                                                order by su.SiparisTarihi desc
                                           
                                           
                                           """,(id))
        
        return topResult[0][0]

class TeklifMusteriler:
    def __init__(self):
        self.data = SqlConnect().data
    
    def getTeklifMusteriler(self):
        try:
            result = self.data.getList("""
                                            select 

                                                *,
                                                (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = ym.UlkeId) as UlkeAdi,
												(select k.KullaniciAdi from KullaniciTB k where k.ID = ym.Kullanici) as KullaniciAdi 

                                            from YeniTeklif_MusterilerTB ym
                                       """)
            liste = list()
            for item in result:
                model = TeklifMusterilerModel()
                model.id = item.Id
                model.customer = item.MusteriAdi
                model.email = item.Mail
                model.company = item.Company
                model.phone = item.Phone
                model.country = item.UlkeId
                model.countryName = item.UlkeAdi
                model.user = item.Kullanici
                model.adress = item.Adress
                model.username = item.KullaniciAdi
                liste.append(model)
                
            schema = TeklifMusterilerSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print("getTeklifMusteriler",str(e))
            return False
    
    def getTeklifMusterilerAyrinti(self,id):
        try:
            result = self.data.getStoreList("""
                                            select * from YeniTeklif_MusterilerTB where Id = ?
                                       """,(id))
            liste = list()
            for item in result:
                model = TeklifMusterilerModel()
                model.id = item.Id
                model.customer = item.MusteriAdi
                model.email = item.Mail
                model.company = item.Company
                model.phone = item.Phone
                model.country = item.UlkeId
                model.user = item.Kullanici
                model.adress = item.Adress
                liste.append(model)
                
            schema = TeklifMusterilerSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print("getTeklifMusterilerAyrinti",str(e))
            return False
    
    def setTeklifMusteriler(self,data):
        try:
            
            self.data.update_insert("""
                                        update YeniTeklif_MusterilerTB SET MusteriAdi=?,Company=?,Mail=?,Phone=?,UlkeId=?,Adress=? where Id=?

                                    
                                    """,(data['customer'],data['company'],data['email'],data['phone'],data['country'],data['adress'],data['id'])) 
            return True
        except Exception as e:
            print("setTeklifMusteriler hata",str(e))
            return False
        
    def setTeklifMusterilerKayit(self,data):
        try:
            
            self.data.update_insert("""
                                        insert into YeniTeklif_MusterilerTB(MusteriAdi,Company,Mail,Phone,UlkeId,Kullanici,Adress) VALUES(?,?,?,?,?,?,?)


                                    
                                    """,(data['customer'],data['company'],data['email'],data['phone'],data['country'],data['user'],data['adress'])) 
            return True
        except Exception as e:
            print("setTeklifMusteriler hata",str(e))
            return False
        
    def setTeklifMusterilerSil(self,id):
        self.data.update_insert("""
                                    delete YeniTeklif_MusterilerTB where Id =?
                                
                                
                                """,(id))
        
        return True      
    
    def setTeklifMusterilerKopyala(self,data):
        try:
            self.data.update_insert("""

                                        insert into 
                                        MusterilerTB(FirmaAdi,
                                        UlkeId,
                                        Unvan,
                                        Adres,
                                        MailAdresi,
                                        Telefon,
                                        Marketing,
                                        MusteriTemsilciId,
                                        KullaniciID,
                                        Ozel,
                                        Satisci,
                                        MusteriOncelik,
                                        Aktif) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                                    
                                    """,(data['customer'],
                                         data['country'],
                                         data['company'],
                                         data['adress'],
                                         data['email'],
                                         data['phone'],
                                         'Mekmar',
                                         data['user'],
                                         data['user'],
                                         1,
                                         data['user'],
                                         'B',
                                         1
                                         
                                         ))
            return True
        except Exception as e:
            print('setTeklifMusterilerKopyala hata',str(e))
        
 
class FuarMusteriler:
    def __init__(self):
        self.data = SqlConnect().data   

    def setFuarMusterilerKayit(self,data):
        try:
            self.data.update_insert("""
                                        insert into FuarMusterilerTB(Customer,Company,Email,Phone,Country,Kullanici,Adress,Orderer,FilelinkOn,FilelinkArka) VALUES(?,?,?,?,?,?,?,?,?,?)
                                    
                                    """,(data['customer'],data['company'],data['email'],data['phone'],data['country'],data['user'],data['adress'],data['orderer'],data['linkOn'],data['linkArka']))
            
            return True
        except Exception as e:
            print("setFuarMusterilerKayit hata",str(e))
            return False

    def setFuarMusterilerGuncelle(self,data):
        try:
            self.data.update_insert("""
                                        update FuarMusterilerTB SET Customer=?,Company=?,Email=?,Phone=?,Country=?,Adress=?,Orderer=?,FilelinkOn=?,FilelinkArka=? WHERE ID=?
                                    
                                    """,(data['customer'],data['company'],data['email'],data['phone'],data['country'],data['adress'],data['satisci'],data['linkOn'],data['linkArka'],data['id']))
            
            return True
        except Exception as e:
            print("setFuarMusterilerKayit hata",str(e))
            return False


    
    def getFuarMusterileriList(self):
        try:
            result = self.data.getList("""
                                        select

*,
(select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = fm.Country) as Ulke


from FuarMusterilerTB fm
                                       
            
            
                                       """)
            liste = list()
            for item in result:
                model = FuarMusterilerModel()
                model.id = item.ID
                model.customer = item.Customer
                model.email = item.Email
                model.company = item.Company
                model.phone = item.Phone
                model.country = item.Ulke
                model.user = item.Kullanici
                model.adress = item.Adress
                model.satisci = item.Orderer
                
                liste.append(model)
            schema = FuarMusterilerSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print("getFuarMusterileri hata ",str(e))
            return False
        

    def getFuarMusterileriListAyrinti(self,id):
        try:
            result = self.data.getStoreList("""
                                        select * from FuarMusterilerTB where ID=?
                                       
            
            
                                       """,(id))
            liste = list()
            for item in result:
                model = FuarMusterilerModel()
                model.id = item.ID
                model.customer = item.Customer
                model.email = item.Email
                model.company = item.Company
                model.phone = item.Phone
                model.country = item.Country
                model.user = item.Kullanici
                model.adress = item.Adress
                model.satisci = item.Orderer
                model.linkOn = item.FilelinkOn
                model.linkArka = item.FilelinkArka
                liste.append(model)
            schema = FuarMusterilerSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print("getFuarMusterileri hata ",str(e))
            return False
        
    def getFuarMusterileriListSil(self,id):
        
        self.data.update_insert("""
                                
                                    delete FuarMusterilerTB where ID =?
                                
                                
                                """,(id))
        return True
   
class BgpMusteriler:
    def __init__(self):
        self.data = SqlConnect().data   

    def setBgpMusterilerKayit(self,data):
        try:
            self.data.update_insert("""
                                        insert into BgpProjectMusteriler(Customer,Company,Email,Phone,Ulke,Kullanici,Adress,Satisci) VALUES(?,?,?,?,?,?,?,?)
                                    
                                    """,(data['customer'],data['company'],data['email'],data['phone'],data['country'],data['user'],data['adress'],data['orderer']))
            
            return True
        except Exception as e:
            print("setBgpMusterilerKayit hata",str(e))
            return False

    def setBgpMusterilerGuncelle(self,data):
        try:
            print(data)
            self.data.update_insert("""
                                        update BgpProjectMusteriler SET Customer=?,Company=?,Email=?,Phone=?,Ulke=?,Adress=?,Satisci=? WHERE ID=?
                                    
                                    """,(data['customer'],data['company'],data['email'],data['phone'],data['country'],data['adress'],data['satisci'],data['id']))
            
            return True
        except Exception as e:
            print("setBgpMusterilerGuncelle hata",str(e))
            return False


    
    def getBgpMusterileriList(self):
        try:
            result = self.data.getList("""
                                        select 

                                        *,
                                        (select KullaniciAdi from KullaniciTB k where k.ID =bgp.Kullanici) as SatisciAdi

                                        from BgpProjectMusteriler bgp
                                       
            
            
                                       """)
            liste = list()
            for item in result:
                model = BgpMusterilerModel()
                model.id = item.ID
                model.customer = item.Customer
                model.email = item.Email
                model.company = item.Company
                model.phone = item.Phone
                model.country = item.Ulke
                model.user = item.Kullanici
                model.adress = item.Adress
                model.satisci = item.SatisciAdi
                
                liste.append(model)
            schema = BgpMusterilerSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print("getBgpMusterileriList hata ",str(e))
            return False
        

    def getBgpMusterileriListAyrinti(self,id):
        try:
            result = self.data.getStoreList("""
                                        select * from BgpProjectMusteriler where ID=?
                                       
            
            
                                       """,(id))
            liste = list()
            for item in result:
                model = BgpMusterilerModel()
                model.id = item.ID
                model.customer = item.Customer
                model.email = item.Email
                model.company = item.Company
                model.phone = item.Phone
                model.country = item.Ulke
                model.user = item.Kullanici
                model.adress = item.Adress
                model.satisci = item.Satisci
                liste.append(model)
            schema = BgpMusterilerSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print("getBgpMusterileriListAyrinti hata ",str(e))
            return False
        
    def getBgpMusterileriListSil(self,id):
        
        self.data.update_insert("""
                                
                                    delete BgpProjectMusteriler where ID =?
                                
                                
                                """,(id))
        return True
   
 


