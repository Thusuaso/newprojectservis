
from helpers import SqlConnect
from openpyxl import *
from openpyxl.styles import PatternFill
from openpyxl.cell import Cell
from openpyxl.styles import Alignment  

from openpyxl.drawing.image import Image
import datetime 
from helpers import SqlConnect,TarihIslemler
import shutil

class ExcelCiktiIslem:

    def uretimCiktiEn(self,data_list):
        try:
            source_path = 'resource_api/siparisler/sablonlar/Uretim_list.xlsx'
            target_path = 'resource_api/siparisler/dosyalar/Uretim_list.xlsx'
            shutil.copy2(source_path, target_path)
          
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')
             
            satir = 2
            item = 0 
            j = len(data_list)-1
            
            m = 0
            a = 0
            k = 0
            #rgb=[241,195,170]
            ##color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
           
            while item <= j  :
               
                 
                
                 sayfa.cell(satir,column=2,value=data_list[item]['tarih']).alignment = Alignment(horizontal='center', vertical='center') 
                 sayfa.cell(satir,column=4,value=data_list[item]['musteriAdi']).alignment = Alignment(horizontal='center', vertical='center')
                 sayfa.cell(satir,column=6,value=data_list[item]['siparisNo']).alignment = Alignment(horizontal='center', vertical='center')    
                 k = satir
                 satir += 1
                 i = 0
               
                 m = 0
                 a=0
                 
                 for item1 in data_list:
                   
                    if data_list[i]['siparisNo'] == data_list[item]['siparisNo']: 
                         sayfa.cell(satir-1,column=1,value=item1['tarih'])
                         sayfa.cell(satir-1,column=3,value=item1['musteriAdi'])
                         sayfa.cell(satir-1,column=5,value=item1['siparisNo'])
                         sayfa.cell(satir-1,column=7,value=item1['urunAdi'])
                         sayfa.cell(satir-1,column=8,value=item1['musteriAciklama'])
                         sayfa.cell(satir-1,column=9,value=item1['en'] + 'x'+ item1['boy']+ 'x'+ item1['kenar'])
                         sayfa.cell(satir-1,column=10,value=item1['tedarikciAdi'])
                         sayfa.cell(satir-1,column=11,value=item1['siparisMiktari'])
                         sayfa.cell(satir-1,column=12,value=item1['uretimMiktari'])
                         sayfa.cell(satir-1,column=13,value=item1['birim'])
                         sayfa.cell(satir-1,column=14,value=item1['birimFiyat'])
                         sayfa.cell(satir-1,column=15,value=item1['satisToplam'])
                         satir += 1
                         a += 1
                      
                         i +=1
                         m +=1
                        
                        
                    else :  i +=1
                 satir = satir-2        
                 sayfa.merge_cells(start_row=k,start_column=2,end_row=satir,end_column=2)##.fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)     
                 sayfa.merge_cells(start_row=k,start_column=4,end_row=satir,end_column=4)
                 sayfa.merge_cells(start_row=k,start_column=6,end_row=satir,end_column=6)
                 item = item + m
                 if i <j : i +=1 
                 satir += 1
                        
                    
            ##sayfa.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
            kitap.save(target_path)
            kitap.close()
            
            return True

        except Exception as e:
            print('ExcelCiktiIslem uretimCikti Hata : ',str(e))
            return False



    def uretimCikti(self,data_list):     
        try:
            source_path = 'resource_api/siparisler/sablonlar/Uretim_list.xlsx'
            target_path = 'resource_api/siparisler/dosyalar/Uretim_list.xlsx'
            shutil.copy2(source_path, target_path)
          
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')
             
            satir = 2
            item = 0 
            j = len(data_list)-1
            
            m = 0
            a = 0
            k = 0
            #rgb=[241,195,170]
            ##color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
           
            while item <= j  :
               
                 
                
                 sayfa.cell(satir,column=2,value=data_list[item]['tarih']).alignment = Alignment(horizontal='center', vertical='center') 
                 sayfa.cell(satir,column=4,value=data_list[item]['musteriAdi']).alignment = Alignment(horizontal='center', vertical='center')
                 sayfa.cell(satir,column=6,value=data_list[item]['siparisNo']).alignment = Alignment(horizontal='center', vertical='center')    
                 k = satir
                 satir += 1
                 i = 0
               
                 m = 0
                 a=0
                 
                 for item1 in data_list:
                   
                    if data_list[i]['siparisNo'] == data_list[item]['siparisNo']: 
                         sayfa.cell(satir-1,column=1,value=item1['tarih'])
                         sayfa.cell(satir-1,column=3,value=item1['musteriAdi'])
                         sayfa.cell(satir-1,column=5,value=item1['siparisNo'])
                         sayfa.cell(satir-1,column=7,value=item1['urunAdi'])
                         sayfa.cell(satir-1,column=8,value=item1['icerik'])
                         sayfa.cell(satir-1,column=9,value=item1['en'] + 'x'+ item1['boy']+ 'x'+ item1['kenar'])
                         sayfa.cell(satir-1,column=10,value=item1['tedarikciAdi'])
                         sayfa.cell(satir-1,column=11,value=item1['siparisMiktari'])
                         sayfa.cell(satir-1,column=12,value=item1['uretimMiktari'])
                         sayfa.cell(satir-1,column=13,value=item1['birim'])
                         sayfa.cell(satir-1,column=14,value=item1['birimFiyat'])
                         sayfa.cell(satir-1,column=15,value=item1['satisToplam'])
                         satir += 1
                         a += 1
                      
                         i +=1
                         m +=1
                        
                        
                    else :  i +=1
                 satir = satir-2        
                 sayfa.merge_cells(start_row=k,start_column=2,end_row=satir,end_column=2)##.fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)     
                 sayfa.merge_cells(start_row=k,start_column=4,end_row=satir,end_column=4)
                 sayfa.merge_cells(start_row=k,start_column=6,end_row=satir,end_column=6)
                 item = item + m
                 if i <j : i +=1 
                 satir += 1
                        
                    
            ##sayfa.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
            kitap.save(target_path)
            kitap.close()
            
            return True

        except Exception as e:
            print('ExcelCiktiIslem uretimCikti Hata : ',str(e))
            return False
     
     


    def IcSiparisExcelCikti(self,data_list):
        tarihIslem = TarihIslemler()
        e = datetime.datetime.now()
        try:
            source_path = 'resource_api/siparisler/sablonlar/İç Sipariş Formu.xlsx'
            target_path = 'resource_api/siparisler/dosyalar/İç Sipariş Formu.xlsx'

            shutil.copy2(source_path, target_path)
            
            
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
          
          
            satir = 17
            no = 1
            i = 7
            item1 = 1
            miktari = 0
            tarih = e
           
            say = data_list[5]
            say = int(say) + 16
           
            satistop = 0
            while item1 <= len(data_list)-7:
                if data_list[i]['adet'] !=0:
                    miktari = data_list[i]['adet']

                elif  data_list[i]['m2'] !=0:   
                    miktari = data_list[i]['m2']
                
                elif  data_list[i]['mt'] !=0:   
                    miktari  = data_list[i]['mt']

                elif  data_list[i]['ton'] !=0:   
                    miktari = data_list[i]['ton']                 
                else : miktari = 0
               
                sayfa.cell(11,column=3,value=data_list[i]['siparisNo'])
                sayfa.cell(12,column=3,value=tarih)
                sayfa.cell(13,column=3,value=data_list[i]['tedarikciAdi'])
                sayfa.cell(11,column=7,value=data_list[1])
                
                sayfa.cell(12,column=7,value=data_list[0]['teslimAdi'])    
                sayfa.cell(13,column=7,value=data_list[2][0]['faturaTur'])          
                
                sayfa.cell(satir,column=1,value=no)
               
                sayfa.cell(satir,column=2,value=data_list[i]['uretimAciklama'])
               
                sayfa.cell(satir,column=3,value=data_list[i]['yuzeyIslem'])
                sayfa.cell(satir,column=4,value=data_list[i]['en']+"x"+data_list[i]['boy']+"x"+data_list[i]['kenar'])
                
                sayfa.cell(satir,column=5,value=data_list[i]['kasaAdet'])
               
              
                sayfa.cell(satir,column=7,value=miktari)
               
                sayfa.cell(satir,column=8,value=data_list[i]['urunbirimAdi'])
                sayfa.cell(satir,column=9,value=data_list[i]['alisFiyati'])
                sayfa.cell(satir,column=10,value=float(data_list[i]['alisFiyati'])*data_list[i]['miktar'])
                satistop += float(data_list[i]['alisFiyati'])*data_list[i]['miktar']
                item1 = item1 + 1
                i +=1
                no +=1
               
                satir += 1
            
          
            sayfa.cell(satir+2,column=10,value=satistop)
            sayfa.cell(33,column=8,value=data_list[0]['teslimAdi'])
            sayfa.cell(33,column=10,value=satistop)
            sayfa.cell(44,column=1,value="4."+data_list[3])
            sayfa.cell(45,column=1,value="5."+data_list[4])
            sayfa.cell(say,column=6, value=data_list[6])
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('IcSiparisExcelCikti depoCikti Hata : ',str(e))
            return False
       
       
       