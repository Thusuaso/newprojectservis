from openpyxl import *
import shutil


class ExcellCiktiIslem:

    def ceki_listesi_excel(self,data_list):
         try:
            source_path = 'resource_api/shared/sablonlar/ceki_listesi.xlsx'
            target_path = 'resource_api/shared/dosyalar/ceki_listesi.xlsx'

            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')

            satir = 11
           
            for item in data_list:

                sayfa.cell(satir,column=2,value=item['kasaNo'])
                sayfa.cell(satir,column=3,value=item['kategoriAdi'])
                sayfa.cell(satir,column=4,value=item['yuzeyIslem'])
                sayfa.cell(satir,column=5,value=item['urunAdi'])
                sayfa.cell(satir,column=6,value=item['kenar'])
                sayfa.cell(satir,column=7,value=item['en'])
                sayfa.cell(satir,column=8,value=item['boy'])
                sayfa.cell(satir,column=9,value=item['adet'])
                

                miktar = 0
                kutu = int(item['kutuAdet'])
                if(item['birimAdi'] == 'M2'):
                    
                    if(item['en']=='ANT' and item['boy']=='PAT'):
                        miktar = float(round((0.74338688 * kutu),2))
                        
                    elif(item['en']=='20,3' and item['boy']=='SET'):
                        miktar=float(round((0.494914 * kutu),2))
                    elif(item['en'] == 'VAR'):
                        miktar = float(item['miktar'])
                    else:
                        miktar = float(item['miktar'])
                elif (item['birimAdi'] == 'Adet'):
                    miktar='-'
                elif (item['birimAdi'] == 'Mt'):
                    miktar=float(item['miktar'])
                sayfa.cell(satir,column=10,value=miktar)
                kg = 0
                
                if(item['kenar']):
                    if (item['kenar']=='VAR'):
                        kenar = 1
                    elif(item['kenar'] == 'Various'):
                        kenar = 1
                    else: 
                        kenar = item['kenar'].replace(',','.')
                        kenar = float(kenar)
                        
                
                else:
                    kenar=1
                if (item['birimAdi'] == 'M2'):
                    if (item['kategoriAdi'] == 'Travertine Tiles'):
                        kg = int(round((kenar * miktar * 10.0 * 2.40),0))
                    elif (item['kategoriAdi'] == 'Marble Tiles'):
                        kg = int(round((kenar * miktar * 10.0 * 2.80),0))
                    elif (item['kategoriAdi'] == 'Other'):
                        kg = int(round((kenar * miktar * 10 * 1),0))
                    elif ((item['kategoriAdi'] == 'Travertine Mosaic') and (item['yuzeyIslem'] == 'Split face')):
                        kg = int(round((1.5 * miktar * 10 * 2.40),0))
                    else:
                        kg=0
                    
                    
                sayfa.cell(satir,column=11,value=kg)
                sayfa.cell(satir,column=12,value=kg + 30)
                sayfa.cell(satir,column=13,value=item['kasaOlcusu'])
                
                
                
                
                satir += 1
          
            kitap.save(target_path)
            kitap.close()

            return True

         except Exception as e:
            print('ceki_listesi_excel  Hata : ',str(e))
            return False  
    
    
    def mk_sevk_sip_rapor_excel(self,data):
        try:
            source_path = 'resource_api/shared/sablonlar/mk_sevk_sip_listesi.xlsx'
            target_path = 'resource_api/shared/dosyalar/mk_sevk_sip_listesi.xlsx'

            shutil.copy2(source_path, target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
           
            for item in data:

                sayfa.cell(satir,column=1,value=item['musteriadi'])
                sayfa.cell(satir,column=2,value=float(item['siparisfob']))
                sayfa.cell(satir,column=3,value=float(item['yuklenenddp']))
                sayfa.cell(satir,column=4,value=float(item['total']))
                
                satir += 1
          
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ceki_listesi_excel  Hata : ',str(e))
            return False     
    
        
        