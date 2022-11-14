from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil


class ExcelCiktiIslem:

    def konteynerCikti(self,data_list):
        try:
            source_path = 'resource_api/efesfinans/ekonteyner_islem/sablonlar/Konteyner.xlsx'
            target_path = 'resource_api/efesfinans/ekonteyner_islem/dosyalar/Konteyner.xlsx'

            shutil.copy2(source_path, target_path)

            rgb=[249, 255, 130]
            color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
           
            pestop = 0 #peşinatın toplamı
            gentop = 0 #genel bakiyenın toplamı
            for item in data_list:

                sayfa.cell(satir,column=1,value=item['musteriadi'])
                sayfa.cell(satir,column=2,value=item['eski_pesinat'])
                sayfa.cell(satir,column=3,value=item['devir'])
                sayfa.cell(satir,column=4,value=item['ciro'])
                sayfa.cell(satir,column=5,value=item['odenen'])
                sayfa.cell(satir,column=6,value=item['bakiye'])
                sayfa.cell(satir,column=7,value=item['pesinat']).fill = PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
                sayfa.cell(satir,column=8,value=item['genel_bakiye']).fill = PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
                pestop += item['pesinat']
                gentop += item['genel_bakiye']
                satir += 1
            sayfa.cell(satir+2,column=7,value=pestop)
            sayfa.cell(satir+2,column=8,value=gentop)
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem konteynerCikti Hata : ',str(e))
            return False

    def tahsilatCikti(self,data_list):

        try:
            source_path = 'resource_api/efesfinans/ekonteyner_islem/sablonlar/Tahsilat.xlsx'
            target_path = 'resource_api/efesfinans/ekonteyner_islem/dosyalar/Tahsilat.xlsx'

            shutil.copy2(source_path, target_path)

            rgb=[249, 255, 130]
            color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
           
            top = 0 #odemelerin toplamı
       
            for item in data_list:

                sayfa.cell(satir,column=1,value=item['tarih'])
                sayfa.cell(satir,column=2,value=item['musteriadi'])
                sayfa.cell(satir,column=3,value=item['siparisno'])
                sayfa.cell(satir,column=4,value=item['tutar']).fill = PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
                top += item['tutar']
                
                satir += 1
            sayfa.cell(satir+2,column=4,value=top)
           
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem konteynerCikti Hata : ',str(e))
            return False


    def konteyner_ayrinti_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/finans/konteyner_islem/sablonlar/konteyner_ayrinti_listesi.xlsx'
            target_path = 'resource_api/finans/konteyner_islem/dosyalar/konteyner_ayrinti_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')

            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['siparisno'])               
                sayfa.cell(satir,column=2,value=item['yuklemetarihi'])
                sayfa.cell(satir,column=3,value=item['tip'])
                sayfa.cell(satir,column=4,value=item['toplam'])
                sayfa.cell(satir,column=5,value=item['kalan'])
                sayfa.cell(satir,column=6,value=item['vade'])


                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False

    def konteyner_odeme_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/finans/konteyner_islem/sablonlar/konteyner_odeme_listesi.xlsx'
            target_path = 'resource_api/finans/konteyner_islem/dosyalar/konteyner_odeme_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['tarih'])               
                sayfa.cell(satir,column=2,value=item['tutar'])

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False

    def musteri_odeme_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/finans/konteyner_islem/sablonlar/musteri_odeme_listesi.xlsx'
            target_path = 'resource_api/finans/konteyner_islem/dosyalar/musteri_odeme_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['tarih'])   
                sayfa.cell(satir,column=2,value=item['musteriadi'])             
                sayfa.cell(satir,column=3,value=item['tutar'])

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False
    