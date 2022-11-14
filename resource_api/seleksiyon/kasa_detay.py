from helpers import SqlConnect,TarihIslemler
from models.seleksiyon.kasa_detay_olculeri import KasaDetayOlculeriSchema, KasaDetayModel,TedarikcilerListSchema,TedarikcilerListModel
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil

class KasaDetayListesi:
    def __init__(self):
        self.data = SqlConnect().data


    def getKasaDetay(self):

        result = self.data.getList(
            """
                select ks.Ebat,ks.KasaOlculeri,td.FirmaAdi,ks.Id from kasa_detay_olculeri as ks, TedarikciTB as td WHERE ks.Tedarikci = td.ID
            """
          
        )

        liste = list()
        for item in result:

            model = KasaDetayModel()
            model.id=item.Id
            model.ebat = item.Ebat 
            model.kasaOlculeri = item.KasaOlculeri
            model.firmaadi = item.FirmaAdi
            
          
            


            liste.append(model)

        schema = KasaDetayOlculeriSchema(many=True)
       
        return schema.dump(liste)

class KasaDetayExcell:
    def kasa_detay_excell(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/kasadetay_listesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/kasadetay_listesi.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 3

            for item in data_list:
               
                sayfa.cell(satir,column=1,value=item['ebat'])  
                sayfa.cell(satir,column=2,value=item['kasaOlculeri'])             
                sayfa.cell(satir,column=3,value=item['firmaadi'])
                sayfa.cell(satir,column=4,value=item['adet'])
               

                satir += 1
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('KasaDetayExcell depoCikti Hata : ',str(e))
            return False
        
class KasaDetayGuncelle:
    def __init__(self):
        self.data = SqlConnect().data
        
    def setKasaDetayGuncelle(self,data):
        try:
            self.data.update_insert("""
                                        update kasa_detay_olculeri SET Ebat=?,KasaOlculeri=?,Tedarikci=? where Id=?
                                    
                                    """,(data['tasEbat'],data['kasaOlcusu'],data['firmaId'],data['id']))
            
            return True
        except Exception as e:
            print('Kasa Detay Güncelleme Hata',str(e))
            
            return False
        
        
class TedarikciList:
    
    def __init__(self):
        self.data = SqlConnect().data
        
    def getTedarikciList(self):
        
        try:
            result = self.data.getList("""
                                            select tb.ID,tb.FirmaAdi from TedarikciTB tb

                                       """) 
            liste = list()
            for item in result:
                model = TedarikcilerListModel()
                model.id = item.ID
                model.firmaAdi = item.FirmaAdi
                liste.append(model)
            
            schema = TedarikcilerListSchema(many=True)
       
            return schema.dump(liste)
            
            
            return result
        except Exception as e:
            print('Tedarikci Listesi Hata',str(e))
            return False
        
class KasaDetayKaydet:
    def __init__(self):
        self.data = SqlConnect().data
        
    def kaydet(self,datas):
        
        try:
            
            self.data.update_insert("""
                                
                                    insert into kasa_detay_olculeri(Ebat,Tedarikci,KasaOlculeri,Adet) VALUES(?,?,?,?)
                                
                                """,(datas['ebat'],datas['firmaId'],datas['kasaEbat'],0))
            return True
        except Exception as e:
            print('Kasa Detay Kaydetme Hata',str(e))
            return False
class KasaDetaySil:
    def __init__(self):
        self.data = SqlConnect().data
        
    def sil(self,id):
        try:
            self.data.update_insert("""
                                        delete kasa_detay_olculeri where Id=?
                                    
                                    
                                    """,(id))
            return True
        except Exception as e:
            print('Kasa Detay Silme Başarısız',str(e))
            return False