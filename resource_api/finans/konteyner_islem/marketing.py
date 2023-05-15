
import datetime
from helpers import SqlConnect
from models.finans import *
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

import shutil
class Marketing:

    def __init__(self):
        self.data = SqlConnect().data
        self.navlunYukleme = []
        self.navlunUretim = []
        self.liste = []
        self.marketingNavlun = []
        self.navlunveDigerleriMonth = self.data.getList("""
                                                    select 


                                                        sum(s.NavlunSatis) +
                                                        sum(s.DetayTutar_1) +
                                                        sum(s.DetayTutar_2) +
                                                        sum(s.DetayTutar_3) +
                                                        sum(s.DetayTutar_4)  as Masraf,


                                                        MONTH(s.YuklemeTarihi) as Month


                                                    from SiparislerTB s
                                                        inner join MusterilerTB m on m.ID = s.MusteriID

                                                    where
                                                        YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and m.Marketing='İç Piyasa'

                                                    group by MONTH(s.YuklemeTarihi)
                                                  
                                                  """)
        self.navlunveDigerleriMonthMekmer = self.data.getList("""
                                                                    select 


                                                                        sum(s.NavlunSatis) +
                                                                        sum(s.DetayTutar_1) +
                                                                        sum(s.DetayTutar_2) +
                                                                        sum(s.DetayTutar_3) +
                                                                        sum(s.DetayTutar_4)  as Masraf,


                                                                        MONTH(s.YuklemeTarihi) as Month


                                                                    from SiparislerTB s
                                                                        inner join MusterilerTB m on m.ID = s.MusteriID

                                                                    where
                                                                        YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and m.Marketing='Mekmer'

                                                                    group by MONTH(s.YuklemeTarihi)
                                                                 
                                                                 
                                                                 """)
        self.icPiyasaAyrintiMasraflar = []
        self.mekmerAyrintiMasraflar = []
        self.poBazindaNavlun = []

    def getMarketingYuklemeHepsi(self):
        result = self.data.getList("""
                                            select 	
                                                sum(su.SatisToplam) as Toplam,
                                                m.Marketing as Marketing
                                            from MusterilerTB m	
                                                inner join SiparislerTB s on s.MusteriID = m.ID
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                            where 	
                                                YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3  and m.Marketing != 'Seleksiyon'
                                            group by
                                                m.Marketing

                                        """)
        
        self.navlunYukleme = self.data.getList("""
                                                select 	
                                                    sum(s.NavlunSatis) as Navlun,
                                                    sum(s.DetayTutar_1) as DetayTutar1,
                                                    sum(s.DetayTutar_2) as DetayTutar2,
                                                    sum(s.DetayTutar_3) as DetayTutar3,
                                                    sum(s.DetayTutar_4) as DetayTutar4,
                                                    
                                                    m.Marketing as Marketing
                                                from MusterilerTB m	
                                                    inner join SiparislerTB s on s.MusteriID = m.ID
                                                where 	
                                                    YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 
                                                group by
                                                    m.Marketing
                                                                                            
                                             
                                             """)
        
        
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = item.Toplam
            model.cfrToplam = item.Toplam + self.__getNavlun(item.Marketing)
            self.liste.append(model)
        # for item in self.__getMarketingYuklemeGhanaHepsi():
        #     self.liste.append(item)
        # for item in self.__getMayaContainerHepsi():
        #     self.liste.append(item)
        
        # for item in self.__getVilloContainerHepsi():
        #     self.liste.append(item)

            
        schema = MarketingSchema(many=True)
        return schema.dump(self.liste)
    
    def getMarketingYukleme(self,year):
        
        result = self.data.getStoreList("""
                                            select 	
                                                sum(su.SatisToplam) as Toplam,
                                                m.Marketing as Marketing
                                            from MusterilerTB m	
                                                inner join SiparislerTB s on s.MusteriID = m.ID
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                            where 	
                                                YEAR(s.YuklemeTarihi) = ? and s.SiparisDurumID =3 
                                            group by
                                                m.Marketing

                                        """,(year))
        
        self.navlunYukleme = self.data.getStoreList("""
                                                select 	
                                                    sum(s.NavlunSatis) as Navlun,
                                                    sum(s.DetayTutar_1) as DetayTutar1,
                                                    sum(s.DetayTutar_2) as DetayTutar2,
                                                    sum(s.DetayTutar_3) as DetayTutar3,
                                                    sum(s.DetayTutar_4) as DetayTutar4,
                                                    
                                                    m.Marketing as Marketing
                                                from MusterilerTB m	
                                                    inner join SiparislerTB s on s.MusteriID = m.ID
                                                where 	
                                                    YEAR(s.YuklemeTarihi) = ? and s.SiparisDurumID =3
                                                group by
                                                    m.Marketing
                                                                                            
                                             
                                             """,(year))
        
        
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = item.Toplam
            model.cfrToplam = item.Toplam + self.__getNavlun(item.Marketing)
            self.liste.append(model)
        # for item in self.__getMarketingYuklemeGhana(month):
        #     self.liste.append(item)
        # for item in self.__getMayaContainer(month):
        #     self.liste.append(item)
        
        # for item in self.__getVilloContainer(month):
        #     self.liste.append(item)

            
        schema = MarketingSchema(many=True)
        return schema.dump(self.liste)
    
    def getMarketingDetail(self,year):
        try:
            result = self.data.getStoreList("""
                                                select            
                                                    m.ID as MusteriId,            
                                                    m.FirmaAdi as MusteriAdi,               
                                                    m.Marketing,
                                                    (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
                                            
                                    
                                                            
                                        (          
                                        Select Sum(u.SatisToplam) from SiparislerTB s, SiparisUrunTB u where s.SiparisNo=u.SiparisNo and s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?
                                            
                                        ) as Toplam 


                                                    from            
                                                    MusterilerTB m
                                            
                                            """,(year))
            self.marketingNavlun = self.data.getStoreList("""
                                                select            
                 m.ID as MusteriId,            
                 m.FirmaAdi as MusteriAdi,               
                 m.Marketing,
                (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
          
 
                           
    (          
     Select Sum(s.NavlunSatis) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?  
          
    ) +
	(          
     Select Sum(s.DetayTutar_1) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?
          
    ) +
	(          
     Select Sum(s.DetayTutar_2) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?  
          
    )+
	(          
     Select Sum(s.DetayTutar_3) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=? 
          
    ) +
	(          
     Select Sum(s.DetayTutar_4) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=?  
          
    ) as Masraflar


                from            
                MusterilerTB m
                                            
                                            """,(year,year,year,year,year))    
        
            liste = list()
            for item in result:
                model = MarketingAyrintiModel()
                if(item.Toplam == None):
                    continue
                else:
                    model.musteri = item.MusteriAdi
                    model.marketing = item.Marketing
                    model.toplamFob = item.Toplam
                    model.toplamCfr = float(item.Toplam) + float(self.__getMarketingDetailNavlun(item.MusteriId))
                    liste.append(model)
            schema = MarketingAyrintiSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getMarketingDetail hata',str(e))
    
    def __getMarketingDetailNavlun(self,musteriId):
        for item in self.marketingNavlun:
            if item.MusteriId != musteriId:
                continue
            else:
                if item.Masraflar != None:
                    return item.Masraflar
                else:
                    return 0
    
    def getMarketingDetailHepsi(self):
        try:
            result = self.data.getList("""
                                                select            
                                                    m.ID as MusteriId,            
                                                    m.FirmaAdi as MusteriAdi,               
                                                    m.Marketing,
                                                    (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
                                            
                                    
                                                            
                                        (          
                                        Select Sum(u.SatisToplam) from SiparislerTB s, SiparisUrunTB u where s.SiparisNo=u.SiparisNo and s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) 
                                            
                                        ) as Toplam 


                                                    from            
                                                    MusterilerTB m
                                            
                                            """)
            self.marketingNavlun = self.data.getList("""
                                                select            
                 m.ID as MusteriId,            
                 m.FirmaAdi as MusteriAdi,               
                 m.Marketing,
                (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
          
 
                           
    (          
     Select Sum(s.NavlunSatis) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) 
          
    ) +
	(          
     Select Sum(s.DetayTutar_1) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate())
          
    ) +
	(          
     Select Sum(s.DetayTutar_2) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate())  
          
    )+
	(          
     Select Sum(s.DetayTutar_3) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate())   
          
    ) +
	(          
     Select Sum(s.DetayTutar_4) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) 
          
    ) as Masraflar


                from            
                MusterilerTB m
                                            
                                            """)    
        
            liste = list()
            for item in result:
                model = MarketingAyrintiModel()
                if(item.Toplam == None):
                    continue
                else:
                    model.musteri = item.MusteriAdi
                    model.marketing = item.Marketing
                    model.toplamFob = item.Toplam
                    model.toplamCfr = float(item.Toplam) + float(self.__getMarketingDetailNavlunHepsi(item.MusteriId))
                    liste.append(model)
            schema = MarketingAyrintiSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getMarketingDetail hata',str(e))
    
    def __getMarketingDetailNavlunHepsi(self,musteriId):
        for item in self.marketingNavlun:
            if item.MusteriId != musteriId:
                continue
            else:
                if item.Masraflar != None:
                    return item.Masraflar
                else:
                    return 0
    
    def getMarketingUretim(self):
        
        result = self.data.getList("""
                                            select 	
                                                sum(su.SatisToplam) as Toplam,
                                                m.Marketing as Marketing
                                            from MusterilerTB m	
                                                inner join SiparislerTB s on s.MusteriID = m.ID
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                            where 	
                                                    s.SiparisDurumID =2 
                                            group by
                                                m.Marketing

                                        """)
        
        self.navlunUretim = self.data.getList("""
                                                select 	
                                                    sum(s.NavlunSatis) + sum(s.DetayTutar_1) + sum(s.DetayTutar_2) + sum(s.DetayTutar_3) + sum(s.DetayTutar_4) as Navlun,
                                                    m.Marketing as Marketing
                                                from MusterilerTB m	
                                                    inner join SiparislerTB s on s.MusteriID = m.ID
                                                where 	
                                                    s.SiparisDurumID=2
                                                group by
                                                    m.Marketing
                                                                                            
                                             
                                             """)
        
        
        liste = list()
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = item.Toplam
            model.cfrToplam = item.Toplam + self.__getNavlunUretim(item.Marketing)
            liste.append(model)
        # for item in self.__getMarketingUretimGhana():
        #     liste.append(item)
        # for item in self.__getMayaContainerUretim():
        #     liste.append(item)
        # for item in self.__getVilloContainerYukleme():
        #     liste.append(item)
        schema = MarketingSchema(many=True)
        return schema.dump(liste)

    def getMarketingDepo(self,month):
        result = self.data.getStoreList("""
                                            select 


                                                sum(yu.Total) as Toplam,
                                                yd.CustomersId,
                                                (select ym.CustomersName from YeniDepoMusterilerTB ym where ym.Id = yd.CustomersId) as Marketing


                                            from 
                                                YeniDepoSatisTB yd
                                                inner join YeniDepoSatisUrunlerTB yu on yu.OrderNo = yd.OrderNo

                                            where 
                                                YEAR(yd.Date) = YEAR(GETDATE()) and MONTH(yd.Date) = ? and Payment=1 and Shipped=1

                                            group by
                                                yd.CustomersId
                                        """,(month))
        
        liste = list()
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = float(item.Toplam)  * 0.70
            model.cfrToplam = item.Toplam
            
            liste.append(model)
        schema = MarketingSchema(many=True)
        return schema.dump(liste)
    
    def getMarketingDepoHepsi(self):
        result = self.data.getList("""
                                            select 


                                                sum(yu.Total) as Toplam,
                                                yd.CustomersId,
                                                (select ym.CustomersName from YeniDepoMusterilerTB ym where ym.Id = yd.CustomersId) as Marketing


                                            from 
                                                YeniDepoSatisTB yd
                                                inner join YeniDepoSatisUrunlerTB yu on yu.OrderNo = yd.OrderNo

                                            where 
                                                YEAR(yd.Date) = YEAR(GETDATE()) and Payment=1 and Shipped=1

                                            group by
                                                yd.CustomersId
                                        """)
        
        liste = list()
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = float(item.Toplam)  * 0.70
            model.cfrToplam = item.Toplam
            
            liste.append(model)
        schema = MarketingSchema(many=True)
        return schema.dump(liste)
       
    def __getNavlun(self,marketing):
        for item in self.navlunYukleme:
            if item.Marketing != marketing:
                continue
            else:
                return item.Navlun + item.DetayTutar1 + item.DetayTutar2 + item.DetayTutar3 + item.DetayTutar4
                  
    def __getNavlunUretim(self,marketing):
        for item in self.navlunUretim:
            if item.Marketing != marketing:
                continue
            else:
                return item.Navlun
               
    def getBdDepoList(self):
        result =self.data.getList("""
                            select 


                                sum(ysu.Total) as Toplam,
                                MONTH(ys.Date) as Ay



                            from YeniDepoSatisTB ys
                            inner join YeniDepoSatisUrunlerTB ysu on ys.OrderNo = ysu.OrderNo
                            where YEAR(ys.Date) = YEAR(GETDATE())  and ys.Shipped=1
                            group by MONTH(ys.Date)
                            order by MONTH(ys.Date)
                            
                          
                          
                          
                          """)
        liste = list()
        for item in result:
            model = BdDepoModel()
            model.ay = self.__getMonth(item.Ay)
            model.cfrToplam = item.Toplam
            model.fobToplam = round(float(item.Toplam) * 0.7, 2)
            liste.append(model)
        
        schema = BdDepoSchema(many=True)
        return schema.dump(liste)
    
    def getYuklemeAylikList(self):
        
        date = datetime.datetime.now()
        month = date.month + 1
        liste = list()
        
        for item in range(1,month):
            result = self.data.getStoreList("""
                                    select  
                                    s.YuklemeTarihi,  
                                    s.SiparisNo,  
                                    m.FirmaAdi as MusteriAdi,  
                                    (select Sum(SatisToplam) from SiparisUrunTB su where su.SiparisNo=s.SiparisNo) as Fob,  
                                    (select Sum(SatisToplam) from SiparisUrunTB su where su.SiparisNo=s.SiparisNo)+  
                                    dbo.Get_SiparisNavlun(s.SiparisNo) as Dtp,  
                                    'Konteyner' as Tur,m.Marketing  
                                    from  
                                    SiparislerTB s,MusterilerTB m  
                                    where Year(YuklemeTarihi)=YEAR(GETDATE())
                                    and Month(YuklemeTarihi)=?
                                    and m.ID=s.MusteriID  
                                    and m.Marketing not in ('Mekmar Numune','Seleksiyon','Warehouse')  
                                    and m.Marketing is not null  
                                    
                                    union  
                                    select  
                                    s.Tarih as YuklemeTarihi,  
                                    s.CikisNo as SiparisNo,  
                                    m.FirmaAdi as MusteriAdi,  
                                    Sum(Toplam) as Fob  
                                    ,Sum((s.BirimFiyat+7.5)*u.Miktar) as Dtp,  
                                    'Depo' as Tur,m.Marketing  
                                    from  
                                    SevkiyatTB s,MusterilerTB m,UretimTB u  
                                    where s.MusteriID=m.ID and u.KasaNo=s.KasaNo  
                                    and Year(s.Tarih)=YEAR(GETDATE()) and Month(s.Tarih)=?
                                    and m.Mt_No=1  
                                    group by  
                                    s.Tarih,s.CikisNo,m.FirmaAdi,m.Marketing 
                                   
                                   """,(item,item))
            model = AylikYuklemeModel() 
            for item2 in result:
                model.ay = self.__getMonth(item)
                model.fobToplam += item2.Fob
                model.cfrToplam += item2.Dtp
            liste.append(model)
            
        schema = AylikYuklemeSchema(many=True)
        return schema.dump(liste)
            
    def __getMonth(self,month):
            monthName = ['', 'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran', 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
            return monthName[month]

    def byMarketingExcellCikti(self,data):
        try:
        
            source_path = r"resource_api/raporlar/sablonlar/by_marketing_excel.xlsx"
            target_path = r"resource_api/raporlar/dosyalar/by_marketing_excel.xlsx"

            shutil.copy2(source_path,target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            border = Border(left=Side(border_style="thin",color='FF000000'),right=Side(border_style="thin",color='FF000000'),top=Side(border_style="thin",color='FF000000'),bottom=Side(border_style="thin",color='FF000000'))
            satir = 3
            satir2 = 3
            satir3 = 3
            toplamMekmarFob = 0
            toplamMekmarCfr = 0
            for item in data['mekmarLoadMonths']:

                sayfa.cell(satir,column=1,value=item['ay']).border = border
                sayfa.cell(satir,column=2,value=item['fobToplam']).border = border
                sayfa.cell(satir,column=3,value=item['cfrToplam']).border = border
                toplamMekmarFob += item['fobToplam']
                toplamMekmarCfr += item['cfrToplam']
                
                satir += 1
            a1 = sayfa.cell(satir,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            a1.border = border
            a2=sayfa.cell(satir,column=2,value=toplamMekmarFob).font = Font(color='f44336',bold=True)
            a2.border = border
            a3=sayfa.cell(satir,column=3,value=toplamMekmarCfr).font = Font(color='f44336',bold=True)
            a3.border = border
            
            
            toplamFobBd= 0
            toplamCfrBd = 0
            for item in data['mekmarBdLoadMonths']:
                sayfa.cell(satir2,column=5,value=item['ay']).border = border
                sayfa.cell(satir2,column=6,value=item['fobToplam']).border = border
                sayfa.cell(satir2,column=7,value=item['cfrToplam']).border = border
                toplamFobBd += item['fobToplam']
                toplamCfrBd += item['cfrToplam']
                satir2 += 1
                
            b1=sayfa.cell(satir2,column=5,value='Toplam').font = Font(color='f44336',bold=True)
            b1.border = border
            b2=sayfa.cell(satir2,column=6,value=toplamFobBd).font = Font(color='f44336',bold=True)
            b2.border = border
            b3=sayfa.cell(satir2,column=7,value=toplamCfrBd).font = Font(color='f44336',bold=True)
            b3.border = border
                
            toplamFobTotal = 0
            toplamCfrTotal = 0
            for item in data['mekmarTotalLoadMonths']:
                sayfa.cell(satir3,column=9,value=item['ay']).border = border
                sayfa.cell(satir3,column=10,value=item['fobToplami']).border = border
                sayfa.cell(satir3,column=11,value=item['cfrToplami']).border = border
                toplamFobTotal += item['fobToplami']
                toplamCfrTotal += item['cfrToplami']
                
                satir3 += 1
                
            c1=sayfa.cell(satir3,column=9,value='Toplam').font = Font(color='f44336',bold=True)
            c1.border = border
            c2=sayfa.cell(satir3,column=10,value=toplamFobTotal).font = Font(color='f44336',bold=True)
            c2.border = border
            c3=sayfa.cell(satir3,column=11,value=toplamCfrTotal).font = Font(color='f44336',bold=True)
            c3.border = border
            
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem byMarketingExcellCikti Hata : ',str(e))
            return False
    
    def byCustomersExcellCikti(self,data):
        try:
            source_path = "resource_api/raporlar/sablonlar/by_customers_excel.xlsx"
            target_path = "resource_api/raporlar/dosyalar/by_customers_excel.xlsx"

            shutil.copy2(source_path,target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
            satir2 = 2
            toplamSatisTotal = 0
            toplamSatisTotalCfr = 0
            border = Border(left=Side(border_style="thin",color='FF000000'),right=Side(border_style="thin",color='FF000000'),top=Side(border_style="thin",color='FF000000'),bottom=Side(border_style="thin",color='FF000000'))
            for item in data['byCustomersProduct']:

                a1 = sayfa.cell(satir,column=1,value=item['musteriAdi']).border = border
                sayfa.cell(satir,column=2,value=item['marketing']).border = border
                sayfa.cell(satir,column=3,value=item['ulkeAdi']).border = border
                
                
                if(item['toplam'] != None):
                    sayfa.cell(satir,column=4,value=item['toplam']).border = border
                else:
                    sayfa.cell(satir,column=4,value=0).border = border
                    
                if(item['toplam'] != None):
                    sayfa.cell(satir,column=5,value=item['toplamCfr']).border = border
                else:
                    sayfa.cell(satir,column=5,value=0).border = border

                if(item['toplam'] != None):
                    toplamSatisTotal += item['toplam']
                    
                if(item['toplamCfr'] != None):
                    toplamSatisTotalCfr += item['toplamCfr']

                

                
                satir += 1
            a1=sayfa.cell(satir,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            
            a7 = sayfa.cell(satir,column=4,value=self.__getNoneType(toplamSatisTotal)).font = Font(color='f44336',bold=True)
            a7.border = border
            a8=sayfa.cell(satir,column=5,value=self.__getNoneType(toplamSatisTotalCfr)).font = Font(color='f44336',bold=True)
            a8.border = border
            
            
            toplamYuklemeFobProduct = 0
            toplamYuklemeCfrProduct = 0
            satir += 2
            sayfa.cell(satir,column=2,value='Marketing Halihazırdaki Üretim')
            for item in data['byMarketingProduct']:

                sayfa.cell(satir,column=1,value=item['marketing']).border=border
                
                if(item['toplam'] != None):
                    sayfa.cell(satir,column=2,value=item['toplam']).border=border
                else:
                    sayfa.cell(satir,column=2,value=0).border=border
                
                if(item['toplamCfr'] != None):
                    sayfa.cell(satir,column=3,value=item['toplamCfr']).border=border
                else:
                    sayfa.cell(satir,column=3,value=0).border=border

                toplamYuklemeFobProduct += self.__getNoneType(item['toplam'])
                toplamYuklemeCfrProduct += self.__getNoneType(item['toplamCfr'])
                
                
                satir += 1
            b1=sayfa.cell(satir,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            b1.border=border
            b2=sayfa.cell(satir,column=2,value=toplamYuklemeFobProduct).font = Font(color='f44336',bold=True)
            b2.border=border
            b3=sayfa.cell(satir,column=3,value=toplamYuklemeCfrProduct).font = Font(color='f44336',bold=True)
            b3.border=border
            
            
            satir2 = 3
            buyilToplam = 0
            toplamFob = 0
            toplamCfr = 0
            for item in data['imperialHomes']:
                sayfa.cell(satir2,column=7,value=item['musteriAdi']).border = border
                sayfa.cell(satir2,column=8,value=self.__getNoneType(item['toplam'])).border = border
                sayfa.cell(satir2,column=9,value=self.__getNoneType(item['toplamCfr'])).border = border
                buyilToplam += self.__getNoneType(item['satisToplamiBuYil'])
                toplamFob += self.__getNoneType(item['toplam'])
                toplamCfr += self.__getNoneType(item['toplamCfr'])
                satir2 += 1
            b1 = sayfa.cell(satir2,column=7,value='Toplam').font = Font(color='f44336',bold=True)
            b1.border = border
            b3=sayfa.cell(satir2,column=8,value=toplamFob).font = Font(color='f44336',bold=True)
            b3.border = border
            b4=sayfa.cell(satir2,column=9,value=toplamCfr).font = Font(color='f44336',bold=True)
            b4.border = border
            
            
            
            satir3 = 3
            buyilToplamMekmar = 0
            toplamFobMekmar = 0
            toplamCfrMekmar = 0
            for item in data['mekmar']:
                sayfa.cell(satir3,column=11,value=item['musteriAdi']).border = border
                sayfa.cell(satir3,column=12,value=self.__getNoneType(item['toplam'])).border = border
                sayfa.cell(satir3,column=13,value=self.__getNoneType(item['toplamCfr'])).border = border
                buyilToplamMekmar += self.__getNoneType(item['satisToplamiBuYil'])
                toplamFobMekmar += self.__getNoneType(item['toplam'])
                toplamCfrMekmar += self.__getNoneType(item['toplamCfr'])
                
                satir3 += 1
            c1=sayfa.cell(satir3,column=11,value='Toplam').font = Font(color='f44336',bold=True)
            c1.border = border
            
            c3=sayfa.cell(satir3,column=12,value=self.__getNoneType(toplamFobMekmar)).font = Font(color='f44336',bold=True)
            c3.border = border
            
            c4=sayfa.cell(satir3,column=13,value=self.__getNoneType(toplamCfrMekmar)).font = Font(color='f44336',bold=True)
            c4.border = border
            
            
            satir4 = 3
            buyilToplamIcPiyasa = 0
            toplamFobIcPiyasa = 0
            toplamCfrIcPiyasa = 0
            for item in data['icPiyasa']:
                sayfa.cell(satir4,column=15,value=item['musteriAdi']).border = border
                sayfa.cell(satir4,column=16,value=self.__getNoneType(item['toplam'])).border = border
                sayfa.cell(satir4,column=17,value=self.__getNoneType(item['toplamCfr'])).border = border
                buyilToplamIcPiyasa += self.__getNoneType(item['satisToplamiBuYil'])
                toplamFobIcPiyasa += self.__getNoneType(item['toplam'])
                toplamCfrIcPiyasa += self.__getNoneType(item['toplamCfr'])
                
                satir4 += 1
            d1=sayfa.cell(satir4,column=15,value='Toplam').font = Font(color='f44336',bold=True)
            d1.border = border
            d3=sayfa.cell(satir4,column=16,value=toplamFobIcPiyasa).font = Font(color='f44336',bold=True)
            d3.border = border
            d4=sayfa.cell(satir4,column=17,value=toplamCfrIcPiyasa).font = Font(color='f44336',bold=True)
            d4.border = border
                

            
            satir5 = 3
            buyilToplamMekmer = 0
            toplamFobMekmer = 0
            toplamCfrMekmer = 0
            for item in data['mekmer']:
                sayfa.cell(satir5,column=19,value=item['musteriAdi']).border = border
                sayfa.cell(satir5,column=20,value=self.__getNoneType(item['toplam'])).border = border
                sayfa.cell(satir5,column=21,value=self.__getNoneType(item['toplamCfr'])).border = border
                buyilToplamMekmer += self.__getNoneType(item['satisToplamiBuYil'])
                toplamFobMekmer += self.__getNoneType(item['toplam'])
                toplamCfrMekmer += self.__getNoneType(item['toplamCfr'])
                
                satir5 += 1
            e1=sayfa.cell(satir5,column=19,value='Toplam').font = Font(color='f44336',bold=True)
            e1.border = border
            
            e3=sayfa.cell(satir5,column=20,value=toplamFobMekmer).font = Font(color='f44336',bold=True)
            e3.border = border
            
            e4=sayfa.cell(satir5,column=21,value=toplamCfrMekmer).font = Font(color='f44336',bold=True)
            e4.border = border

                
            
            

            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem byMarketingExcellCikti Hata : ',str(e))
            return False
    
    def byMarketingDetailExcellCikti(self,data):
        try:
            source_path = "resource_api/raporlar/sablonlar/by_marketing_detail_excel.xlsx"
            target_path = "resource_api/raporlar/dosyalar/by_marketing_detail_excel.xlsx"

            shutil.copy2(source_path,target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 3
            toplamYuklemeFob = 0
            toplamYuklemeCfr = 0
            border = Border(left=Side(border_style="thin",color='FF000000'),right=Side(border_style="thin",color='FF000000'),top=Side(border_style="thin",color='FF000000'),bottom=Side(border_style="thin",color='FF000000'))
            for item in data['byMarketingLoadMonth']:

                sayfa.cell(satir,column=1,value=item['marketing']).border=border
                sayfa.cell(satir,column=2,value=self.__getNoneType(item['fobToplam'])).border=border
                sayfa.cell(satir,column=3,value=self.__getNoneType(item['cfrToplam'])).border=border

                toplamYuklemeFob += self.__getNoneType(item['fobToplam'])
                toplamYuklemeCfr += self.__getNoneType(item['cfrToplam'])
                
                
                satir += 1
            a1=sayfa.cell(satir,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            a1.border=border
            a2=sayfa.cell(satir,column=2,value=toplamYuklemeFob).font = Font(color='f44336',bold=True)
            a2.border=border
            a3=sayfa.cell(satir,column=3,value=toplamYuklemeCfr).font = Font(color='f44336',bold=True)
            a3.border=border
            
            
            toplamYuklemeFobImperialHomes = 0
            toplamYuklemeCfrImperialHomes = 0
            satir4 = 3
            if(len(data['imperialHomes'])>0):
                
                for item in data['imperialHomes']:

                    sayfa.cell(satir4,column=5,value=item['musteri']).border=border
                    
                    sayfa.cell(satir4,column=6,value=self.__getNoneType(item['toplamFob'])).border=border
                    
                    sayfa.cell(satir4,column=7,value=self.__getNoneType(item['toplamCfr'])).border=border

                    toplamYuklemeFobImperialHomes += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrImperialHomes += self.__getNoneType(item['toplamCfr'])
                    
                    
                    satir4 += 1
                d1=sayfa.cell(satir4,column=5,value='Toplam').font = Font(color='f44336',bold=True)
                d1.border=border
                d2=sayfa.cell(satir4,column=6,value=toplamYuklemeCfrImperialHomes).font = Font(color='f44336',bold=True)
                d2.border=border
                d3=sayfa.cell(satir4,column=7,value=toplamYuklemeCfrImperialHomes).font = Font(color='f44336',bold=True)
                d3.border=border
            
            toplamYuklemeFobMekmar = 0
            toplamYuklemeCfrMekmar = 0
            satir5 = 3
            if(len(data['mekmar'])>0):
                
                for item in data['mekmar']:

                    sayfa.cell(satir5,column=9,value=item['musteri']).border=border
                    sayfa.cell(satir5,column=10,value=self.__getNoneType(item['toplamFob'])).border=border
                    sayfa.cell(satir5,column=11,value=self.__getNoneType(item['toplamCfr'])).border=border

                    toplamYuklemeFobMekmar += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrMekmar += self.__getNoneType(item['toplamCfr'])
                    
                    
                    satir5 += 1
                e1=sayfa.cell(satir5,column=9,value='Toplam').font = Font(color='f44336',bold=True)
                e1.border=border
                e2=sayfa.cell(satir5,column=10,value=float(toplamYuklemeFobMekmar)).font = Font(color='f44336',bold=True)
                e2.border=border
                
                e3=sayfa.cell(satir5,column=11,value=float(toplamYuklemeCfrMekmar)).font = Font(color='f44336',bold=True)
                e3.border=border
                
            
            toplamYuklemeFobMekmer = 0
            toplamYuklemeCfrMekmer = 0
            satir6 = 3
            if(len(data['mekmer'])>0):
                
                for item in data['mekmer']:

                    sayfa.cell(satir6,column=13,value=item['musteri']).border=border
                    sayfa.cell(satir6,column=14,value=self.__getNoneType(item['toplamFob'])).border=border                    
                    sayfa.cell(satir6,column=15,value=self.__getNoneType(item['toplamCfr'])).border=border

                    toplamYuklemeFobMekmer += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrMekmer += self.__getNoneType(item['toplamCfr'])
                    
                    
                    satir6 += 1
                g1 = sayfa.cell(satir6,column=13,value='Toplam').font = Font(color='f44336',bold=True)
                g1.border=border
                g2= sayfa.cell(satir6,column=14,value=float(toplamYuklemeFobMekmer)).font = Font(color='f44336',bold=True)
                g2.border=border
                g3=sayfa.cell(satir6,column=15,value=float(toplamYuklemeCfrMekmer)).font = Font(color='f44336',bold=True)
                g3.border=border
               
            # toplamYuklemeFobEfes = 0
            # toplamYuklemeCfrEfes = 0
            # satir7 = 3
            # if(len(data['efes'])>0):
                
            #     for item in data['efes']:

            #         sayfa2.cell(satir7,column=13,value=item['musteri'])
                    
            #         if(item['toplamFob'] != None):
            #             sayfa2.cell(satir7,column=14,value=item['toplamFob'])
            #         else:
            #             sayfa2.cell(satir7,column=14,value=0)
                    
            #         if(item['toplamCfr'] != None):
            #             sayfa2.cell(satir7,column=15,value=item['toplamCfr'])
            #         else:
            #             sayfa2.cell(satir7,column=15,value=0)

            #         toplamYuklemeFobEfes += self.__getNoneType(item['toplamFob'])
            #         toplamYuklemeCfrEfes += self.__getNoneType(item['toplamCfr'])
                    
                    
            #         satir7 += 1
            #     sayfa2.cell(satir7,column=13,value='Toplam').font = Font(color='f44336',bold=True)
            #     sayfa2.cell(satir7,column=14,value=toplamYuklemeFobMekmer).font = Font(color='f44336',bold=True)
            #     sayfa2.cell(satir7,column=15,value=toplamYuklemeCfrMekmer).font = Font(color='f44336',bold=True)
            
            
            toplamYuklemeFobIcPiyasa = 0
            toplamYuklemeCfrIcPiyasa = 0
            satir8 = 3
            if(len(data['icPiyasa'])>0):
                
                for item in data['icPiyasa']:

                    sayfa.cell(satir8,column=17,value=item['musteri']).border=border
                    sayfa.cell(satir8,column=18,value=self.__getNoneType(item['toplamFob'])).border=border
                    sayfa.cell(satir8,column=19,value=self.__getNoneType(item['toplamCfr'])).border=border

                    toplamYuklemeFobIcPiyasa += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrIcPiyasa += self.__getNoneType(item['toplamCfr'])
                    
                    satir8 += 1
                h1 = sayfa.cell(satir8,column=17,value='Toplam').font = Font(color='f44336',bold=True)
                h1.border=border
                h2=sayfa.cell(satir8,column=18,value=float(toplamYuklemeFobIcPiyasa)).font = Font(color='f44336',bold=True)
                h2.border=border
                h3=sayfa.cell(satir8,column=19,value=float(toplamYuklemeCfrIcPiyasa)).font = Font(color='f44336',bold=True)
                h3.border=border
               
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem byMarketingExcellCikti Hata : ',str(e))
            return False
    
    def __getNoneType(self,value):
        if value == None:
            return 0
        else:
            return float(value)
        
    def getMarketingMonthIcPiyasaLoad(self):
        try:
            result = self.data.getList("""
                                            select 
                                                sum(su.SatisToplam) as Fob,
                                                MONTH(s.YuklemeTarihi) as Month
                                            from SiparislerTB s
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                                inner join MusterilerTB m on m.ID = s.MusteriID
                                            where
                                                YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and m.Marketing='İç Piyasa'
                                            group by MONTH(s.YuklemeTarihi)
                                       """)
            liste = list()
            for item in result:
                model = MarketingMonthLoadModel()
                model.month = self.__getMonth(item.Month)
                model.monthNum = item.Month
                model.fob = item.Fob
                model.ddp = item.Fob + self.__getMarketingMonthIcPiyasaMasraf(item.Month)
                liste.append(model)
            schema = MarketingMonthLoadSchema(many=True)
            return schema.dump(liste)
        
        except Exception as e:
            print("getMarketingMonthLoad hata",str(e))
            
    def __getMarketingMonthIcPiyasaMasraf(self,month):
        for item in self.navlunveDigerleriMonth:
            if item.Month == month:
                return self.__getNoneType(item.Masraf)
    
    def getMarketingMonthMekmerLoad(self):
        try:
            result = self.data.getList("""
                                            select 
                                                sum(su.SatisToplam) as Fob,
                                                MONTH(s.YuklemeTarihi) as Month
                                            from SiparislerTB s
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                                inner join MusterilerTB m on m.ID = s.MusteriID
                                            where
                                                YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and m.Marketing='Mekmer'
                                            group by MONTH(s.YuklemeTarihi)
                                       """)
            liste = list()
            for item in result:
                model = MarketingMonthLoadModel()
                model.month = self.__getMonth(item.Month)
                model.monthNum =item.Month
                model.fob = item.Fob
                model.ddp = item.Fob + self.__getMarketingMonthMekmerMasraf(item.Month)
                liste.append(model)
            schema = MarketingMonthLoadSchema(many=True)
            
            return schema.dump(liste)
        
        except Exception as e:
            print("getMarketingMonthMekmerLoad hata",str(e))
            
    def __getMarketingMonthMekmerMasraf(self,month):
        for item in self.navlunveDigerleriMonthMekmer:
            if item.Month == month:
                return self.__getNoneType(item.Masraf)
         
    def getMarketingMonthIcPiyasaAyrintiLoad(self,month):
        result = self.data.getStoreList("""
                                            select 


                                                sum(su.SatisToplam) as Fob,
                                                s.SiparisNo


                                            from SiparislerTB s 
                                                inner join SiparisUrunTB su on s.SiparisNo=su.SiparisNo
                                                inner join MusterilerTB m on m.ID=s.MusteriID


                                            where YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and MONTH(s.YuklemeTarihi) = ? and m.Marketing='İç Piyasa'

                                            group by s.SiparisNo
                                        """,(month))
        self.icPiyasaAyrintiMasraflar = self.data.getStoreList("""
                                                                    select 


                                                                    s.NavlunSatis as Navlun,
                                                                    s.DetayTutar_1 as DetayTutar1,
                                                                    s.DetayTutar_2 as DetayTutar2,
                                                                    s.DetayTutar_3 as DetayTutar3,
                                                                    s.DetayTutar_4 as DetayTutar4,
                                                                    s.SiparisNo


                                                                from SiparislerTB s 
                                                                    inner join MusterilerTB m on m.ID=s.MusteriID


                                                                where YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and MONTH(s.YuklemeTarihi) = ? and m.Marketing='İç Piyasa'
                                                               """,(month))
        liste = list()
        for item in result:
            model = MarketingMonthLoadAyrintiModel()
            model.siparisNo = item.SiparisNo
            model.fob = item.Fob
            navlun,detay1,detay2,detay3,detay4 = self.__getMarketingMonthIcPiyasaAyrintiLoadMasraf(item.SiparisNo)
            model.navlun = navlun
            model.detay1 = detay1
            model.detay2 = detay2
            model.detay3 = detay3
            model.detay4 = detay4
            model.ddp = item.Fob + navlun + detay1 + detay2 + detay3 + detay4
            liste.append(model)
        schema = MarketingMonthLoadAyrintiSchema(many=True)
        return schema.dump(liste)
    
    def __getMarketingMonthIcPiyasaAyrintiLoadMasraf(self,siparisNo):
        for item in self.icPiyasaAyrintiMasraflar:
            if item.SiparisNo == siparisNo:
                return self.__getNoneType(item.Navlun),self.__getNoneType(item.DetayTutar1),self.__getNoneType(item.DetayTutar2),self.__getNoneType(item.DetayTutar3),self.__getNoneType(item.DetayTutar4)
            
    def getMarketingMonthMekmerAyrintiLoad(self,month):
        result = self.data.getStoreList("""
                                            select 


                                                sum(su.SatisToplam) as Fob,
                                                s.SiparisNo


                                            from SiparislerTB s 
                                                inner join SiparisUrunTB su on s.SiparisNo=su.SiparisNo
                                                inner join MusterilerTB m on m.ID=s.MusteriID


                                            where YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and MONTH(s.YuklemeTarihi) = ? and m.Marketing='Mekmer'

                                            group by s.SiparisNo
                                        """,(month))
        self.mekmerAyrintiMasraflar = self.data.getStoreList("""
                                                                    select 


                                                                    s.NavlunSatis as Navlun,
                                                                    s.DetayTutar_1 as Detay1,
                                                                    s.DetayTutar_2  as Detay2,
                                                                    s.DetayTutar_3 as Detay3,
                                                                    s.DetayTutar_4 as Detay4,
                                                                    s.SiparisNo


                                                                from SiparislerTB s 
                                                                    inner join MusterilerTB m on m.ID=s.MusteriID


                                                                where YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and MONTH(s.YuklemeTarihi) = ? and m.Marketing='Mekmer'
                                                               """,(month))
        liste = list()
        for item in result:
            model = MarketingMonthLoadAyrintiModel()
            model.siparisNo = item.SiparisNo
            model.fob = item.Fob
            navlun,detay1,detay2,detay3,detay4 = self.__getMarketingMonthMekmerAyrintiLoadMasraf(item.SiparisNo)
            model.navlun = navlun
            model.detay1 = detay1
            model.detay2 = detay2
            model.detay3 = detay3
            model.detay4 = detay4
            model.ddp = item.Fob + navlun + detay1 + detay2 + detay3 + detay4
            liste.append(model)
        schema = MarketingMonthLoadAyrintiSchema(many=True)
        return schema.dump(liste)
    
    def __getMarketingMonthMekmerAyrintiLoadMasraf(self,siparisNo):
        for item in self.mekmerAyrintiMasraflar:
            if item.SiparisNo == siparisNo:
                return self.__getNoneType(item.Navlun),self.__getNoneType(item.Detay1),self.__getNoneType(item.Detay2),self.__getNoneType(item.Detay3),self.__getNoneType(item.Detay4)
    
    def getPoBazindaYillikSiparisler(self,yil):
        try:
            result = self.data.getStoreList("""
                                                select 
                                                    s.SiparisNo,
                                                    m.FirmaAdi,
                                                    sum(su.SatisToplam) as SiparisToplam,
													st.TeslimTur,
                                                    s.SiparisTarihi
                                                from SiparislerTB s
                                                    inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                                    inner join MusterilerTB m on m.ID = s.MusteriID
													inner join SiparisTeslimTurTB st on st.ID = s.TeslimTurID
                                                where YEAR(s.SiparisTarihi)= ? and m.Marketing = 'Mekmar'

                                                group by
                                                    s.SiparisNo,m.FirmaAdi,st.TeslimTur,s.SiparisTarihi
                                                order by sum(su.SatisToplam) desc
                                            """,(yil))
            self.poBazindaNavlun = self.data.getStoreList("""
                                                            select 
                                                                s.SiparisNo,
                                                                m.FirmaAdi,
                                                                s.NavlunSatis + s.DetayTutar_1 + s.DetayTutar_2 + s.DetayTutar_3 + s.DetayTutar_4 as GelenTotal

                                                            from SiparislerTB s
                                                                inner join MusterilerTB m on m.ID = s.MusteriID
                                                            where YEAR(s.SiparisTarihi)= ? and m.Marketing = 'Mekmar'
                                                          
                                                          """,(yil))
            liste = list()
            for item in result:
                model = PoBazindaYillikModel()
                model.po = item.SiparisNo
                model.fob = self.__getNoneType(item.SiparisToplam)
                model.ddp = model.fob + self.__getNoneType(self.__getPoBazindaNavlun(item.SiparisNo))
                model.teslim = item.TeslimTur
                model.firma = item.FirmaAdi
                model.tarih = item.SiparisTarihi
                liste.append(model)
            schema = PoBazindaYillikSchema(many=True)
            return schema.dump(liste)
            
        except Exception as e:
            print('getPoBazindaYillikSiparisler hata',str(e))
            return False
        
    def __getPoBazindaNavlun(self,po):
        for item in self.poBazindaNavlun:
            if(item.SiparisNo == po):
                return item.GelenTotal
            
    def getPoBazindaYillikSiparislerExcel(self,data):
        try:
            source_path = "resource_api/raporlar/sablonlar/by_po_uretim_liste.xlsx"
            target_path = "resource_api/raporlar/dosyalar/by_po_uretim_liste.xlsx"

            shutil.copy2(source_path,target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
            toplamYuklemeFob = 0
            toplamYuklemeCfr = 0
            border = Border(left=Side(border_style="thin",color='FF000000'),right=Side(border_style="thin",color='FF000000'),top=Side(border_style="thin",color='FF000000'),bottom=Side(border_style="thin",color='FF000000'))
            for item in data:
                sayfa.cell(satir,column=1,value=item['tarih']).border=border
                sayfa.cell(satir,column=2,value=item['firma']).border=border
                sayfa.cell(satir,column=3,value=item['po']).border=border
                sayfa.cell(satir,column=4,value=item['teslim']).border=border
                
                
                if(item['fob'] != None):
                    sayfa.cell(satir,column=5,value=item['fob']).border=border
                else:
                    sayfa.cell(satir,column=5,value=0).border=border
                
                if(item['ddp'] != None):
                    sayfa.cell(satir,column=6,value=item['ddp']).border=border
                else:
                    sayfa.cell(satir,column=6,value=0).border=border

                toplamYuklemeFob += self.__getNoneType(item['fob'])
                toplamYuklemeCfr += self.__getNoneType(item['ddp'])
                
                
                satir += 1
            a1=sayfa.cell(satir,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            a1.border=border
            a2=sayfa.cell(satir,column=5,value=toplamYuklemeFob).font = Font(color='f44336',bold=True)
            a2.border=border
            a3=sayfa.cell(satir,column=6,value=toplamYuklemeCfr).font = Font(color='f44336',bold=True)
            a3.border=border
            

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem byMarketingExcellCikti Hata : ',str(e))
            return False