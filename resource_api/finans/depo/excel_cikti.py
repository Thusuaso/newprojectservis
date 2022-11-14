from openpyxl import *
import shutil


class ExcelCiktiIslem2:

    def getExcelAyrintiList(self,data_list):

        try:
        
            source_path = r"resource_api/finans/depo/sablonlar/depo_ayrinti_listesi.xlsx"
            target_path = r"resource_api/finans/depo/dosyalar/depo_ayrinti_listesi.xlsx"

            shutil.copy2(source_path,target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')

            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['orderno'])
                sayfa.cell(satir,column=2,value=item['tarih'])
                sayfa.cell(satir,column=3,value=item['sevktarihi'])
                sayfa.cell(satir,column=4,value=item['odemetarihi'])
                sayfa.cell(satir,column=5,value=item['status'])
                sayfa.cell(satir,column=6,value=item['notlar'])
                sayfa.cell(satir,column=7,value=item['toplam'])
                sayfa.cell(satir,column=8,value=item['odenen'])
                sayfa.cell(satir,column=9,value=item['bakiye'])

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem konteynerCikti Hata : ',str(e))
            return False

    