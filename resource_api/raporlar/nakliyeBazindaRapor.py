from helpers import SqlConnect,TarihIslemler
from models.raporlar import *
from marshmallow import Schema,fields
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil

class NakliyeBazinda():
    def __init__(self):
        self.data = SqlConnect().data
    def getNakliyeBazinda(self,year):
        try:
            result = self.data.getStoreList("""
                                                select 
                                                    sum(n.Tutar) as Tl,
                                                    sum(sf.Tutar) as Dolar,
                                                    (select f.FirmaAdi from FirmalarTB f where f.ID=n.FirmaID ) as FirmaAdi
                                                from 
                                                    NakliyeFaturaKayitTB n ,SiparisFaturaKayitTB sf
                                                where
                                                    n.FaturaNo+'.pdf' = sf.EvrakAdi
                                                    and YEAR(n.Tarih)=?
                                                group by 
                                                    n.FirmaID
                                                order by 
                                                    Dolar desc
                                            """,year)
            liste = list()
            for item in result:
                model = NakliyeBazindaModel()
                model.firma = item.FirmaAdi
                model.tl = item.Tl
                model.dolar = item.Dolar
                liste.append(model)
            schema = NakliyeBazindaSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getNakliyeBazinda Hata',str(e))
            
    def getNakliyeExcel(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/nakliyeciRapor.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/nakliyeciRapor.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')    
            satir = 2

            for item in data:

                sayfa.cell(satir,column=1,value=item['firma'])
                sayfa.cell(satir,column=2,value=item['dolar'])
                sayfa.cell(satir,column=3,value=item['tl'])
                satir += 1
                
        
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('getNakliyeExcel  Hata : ',str(e))
            return False
            
class NakliyeBazindaSchema(Schema):
    firma = fields.String()
    tl = fields.Float()
    dolar = fields.Float()
    
class NakliyeBazindaModel:
    firma = ""
    tl = 0
    dolar = 0
    