from helpers import SqlConnect,TarihIslemler
from models.raporlar import *
from marshmallow import Schema,fields
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil

class KomisyonMasraflar():
    def __init__(self):
        self.data = SqlConnect().data
    def getKomisyonMasraflar(self,year):
        try:
            result = self.data.getStoreList("""
                                                select 

                                                    sum(s.Komisyon) as Komisyon,
                                                    (select m.FirmaAdi from MusterilerTB m where m.ID = s.MusteriID) as Musteri


                                                from SiparislerTB s

                                                where s.Komisyon > 0 and YEAR(s.YuklemeTarihi)=?

                                                group by
                                                    s.MusteriID
                                                order by
                                                    Komisyon desc
                                            """,year)

            liste = list()
            for item in result:
                model = KomisyonMasraflarModel()
                model.musteri = item.Musteri
                model.komisyon = item.Komisyon
                liste.append(model)
                
            schema = KomisyonMasraflarSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getKomisyonMasraflar Hata',str(e))
            
    def getKomisyonMasraflarExcel(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/komisyonMasraflar.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/komisyonMasraflar.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')    
            satir = 2

            for item in data:

                sayfa.cell(satir,column=1,value=item['musteri'])
                sayfa.cell(satir,column=2,value=item['komisyon'])
            
                satir += 1
                
        
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('getKomisyonMasraflarExcel  Hata : ',str(e))
            return False
            
class KomisyonMasraflarSchema(Schema):
    musteri = fields.String()
    komisyon = fields.Float()
    
class KomisyonMasraflarModel:
    musteri = ""
    komisyon = 0

    