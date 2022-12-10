from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil



class ExcelCiktiIslemMusteri:

    def musteri_rapor_ciktisi(self,data_list):
        
        try:
            source_path = 'resource_api/raporlar/sablonlar/musteri_listesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/musteri_listesi.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:
                sayfa.cell(satir,column=1,value=item['musteri'])
                sayfa.cell(satir,column=2,value=item['ulkeAdi'])
                sayfa.cell(satir,column=3,value=item['temsilci'])
                
                if item['Toplam'] == None:
                    sayfa.cell(satir,column=4,value=0)
                else:
                    sayfa.cell(satir,column=4,value=item['Toplam'])
                
                
                
                if item['BuYilUretim'] == None:
                    sayfa.cell(satir,column=5,value=0) 
                else:
                    sayfa.cell(satir,column=5,value=item['BuYilUretim']) 

                if item['BuYilSevkiyat'] == None:
                    sayfa.cell(satir,column=6,value=0) 
                else:
                    sayfa.cell(satir,column=6,value=item['BuYilSevkiyat']) 
                
                

                if item['GecenYil'] == None:
                    sayfa.cell(satir,column=7,value=0) 
                else:
                    sayfa.cell(satir,column=7,value=item['GecenYil']) 

                if item['OncekiYil'] == None:
                    sayfa.cell(satir,column=8,value=0) 
                else:
                    sayfa.cell(satir,column=8,value=item['OncekiYil'])
                    
                if item['OnDokuzYili'] == None:
                    sayfa.cell(satir,column=9,value=0) 
                else:
                    sayfa.cell(satir,column=9,value=item['OnDokuzYili']) 
                    
                if item['OnSekizYili'] == None:
                    sayfa.cell(satir,column=10,value=0) 
                else:
                    sayfa.cell(satir,column=10,value=item['OnSekizYili'])                 
                
                if item['OnYediYili'] == None:
                    sayfa.cell(satir,column=11,value=0) 
                else:
                    sayfa.cell(satir,column=11,value=item['OnYediYili']) 
                
                if item['OnAltiYili'] == None:
                    sayfa.cell(satir,column=12,value=0) 
                else:
                    sayfa.cell(satir,column=12,value=item['OnAltiYili']) 
                    
                if item['OnBesYili'] == None:
                    sayfa.cell(satir,column=13,value=0) 
                else:
                    sayfa.cell(satir,column=13,value=item['OnBesYili']) 
                
                
                if item['OnDortYili'] == None:
                    sayfa.cell(satir,column=14,value=0) 
                else:
                    sayfa.cell(satir,column=14,value=item['OnDortYili']) 
                
                if item['OnUcYili'] == None:
                    sayfa.cell(satir,column=15,value=0) 
                else:
                    sayfa.cell(satir,column=15,value=item['OnUcYili']) 
                
                if item['OnUcYiliOncesi'] == None:
                    sayfa.cell(satir,column=16,value=0) 
                else:
                    sayfa.cell(satir,column=16,value=item['OnUcYiliOncesi']) 
                
                
                
             

                satir += 1
            
            kitap.save(target_path)
            kitap.close()
            

            return True

        except Exception as e:
            print('MusteriBazındaListe ExcellÇıktı Hata : ',str(e))
            return False
    def musteri_sip_rapor_ciktisi(self,data_list):
        
        
        try:
            source_path = 'resource_api/raporlar/sablonlar/musteriSipYilListesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/musteriSipYilListesi.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:
                sayfa.cell(satir,column=1,value=item['ay'])
                sayfa.cell(satir,column=2,value=item['fob'])
                sayfa.cell(satir,column=3,value=item['ddp'])
                sayfa.cell(satir,column=4,value=item['fark'])
                
                
                
                
             

                satir += 1
            
            kitap.save(target_path)
            kitap.close()
            

            return True

        except Exception as e:
            
            print('MusteriSipYilBazındaListe ExcellÇıktı Hata : ',str(e))
            return False
        
        
        
    def musteri_sip_buYil_ayrinti_rapor_ciktisi(self,data_list):
        
        try:
            source_path = 'resource_api/raporlar/sablonlar/musteriSipBuYilAyrintiListesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/musteriSipBuYilAyrintiListesi.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:
                sayfa.cell(satir,column=1,value=item['musteri'])
                sayfa.cell(satir,column=2,value=item['fob'])
                sayfa.cell(satir,column=3,value=item['navlun'])
                sayfa.cell(satir,column=4,value=item['detay1'])
                sayfa.cell(satir,column=5,value=item['detay2'])
                sayfa.cell(satir,column=6,value=item['detay3'])
                sayfa.cell(satir,column=7,value=item['detay4'])
                sayfa.cell(satir,column=8,value=item['ddp'])
             

                satir += 1
            
            kitap.save(target_path)
            kitap.close()
            

            return True

        except Exception as e:
            
            print('MusteriSipYilBazındaListe ExcellÇıktı Hata : ',str(e))
            return False
        
    def musteri_sip_gecenYil_ayrinti_rapor_ciktisi(self,data_list):
        
        
        
        try:
            source_path = 'resource_api/raporlar/sablonlar/musteriSipGecenYilAyrintiListesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/musteriSipGecenYilAyrintiListesi.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:
                sayfa.cell(satir,column=1,value=item['musteri'])
                sayfa.cell(satir,column=2,value=item['fob'])
                sayfa.cell(satir,column=3,value=item['navlun'])
                sayfa.cell(satir,column=4,value=item['detay1'])
                sayfa.cell(satir,column=5,value=item['detay2'])
                sayfa.cell(satir,column=6,value=item['detay3'])
                sayfa.cell(satir,column=7,value=item['detay4'])
                sayfa.cell(satir,column=8,value=item['ddp'])
             

                satir += 1
            
            kitap.save(target_path)
            kitap.close()
            

            return True

        except Exception as e:
            
            print('MusteriSipGecenYilBazındaListe ExcellÇıktı Hata : ',str(e))
            return False
        
    def musteri_sip_oncekiYil_ayrinti_rapor_ciktisi(self,data_list):
        
        
        
        try:
            source_path = 'resource_api/raporlar/sablonlar/musteriSipOncekiYilAyrintiListesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/musteriSipOncekiYilAyrintiListesi.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:
                sayfa.cell(satir,column=1,value=item['musteri'])
                sayfa.cell(satir,column=2,value=item['fob'])
                sayfa.cell(satir,column=3,value=item['navlun'])
                sayfa.cell(satir,column=4,value=item['detay1'])
                sayfa.cell(satir,column=5,value=item['detay2'])
                sayfa.cell(satir,column=6,value=item['detay3'])
                sayfa.cell(satir,column=7,value=item['detay4'])
                sayfa.cell(satir,column=8,value=item['ddp'])
             

                satir += 1
            
            kitap.save(target_path)
            kitap.close()
            

            return True

        except Exception as e:
            
            print('MusteriSipOncekiYilBazındaListe ExcellÇıktı Hata : ',str(e))
            return False
        
        
        
    def get_ulke_bazinda_sip_top(self,data_list):
        
        
        try:
            source_path = 'resource_api/raporlar/sablonlar/ulkebzindaSevkiyat.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/ulkebzindaSevkiyat.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:
                sayfa.cell(satir,column=1,value=item['ulkeadi'])
                sayfa.cell(satir,column=2,value=item['toplamsevkiyat'])
 
                
                
                
                
             

                satir += 1
            
            kitap.save(target_path)
            kitap.close()
            

            return True

        except Exception as e:
            
            print('Ulke Bazinda Sevkiyat ExcellÇıktı Hata : ',str(e))
            return False
        
        
    def get_month_marketing_excel_cikti(self,data_list):
        
        try:
            source_path = 'resource_api/raporlar/sablonlar/monthMarketingExcel.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/monthMarketingExcel.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            fill = PatternFill(fill_type='solid',start_color='FFFFFFFF',end_color='FF000000')
            satir = 2
            satir2 = 2
            icPiyasaFobToplam = 0
            icPiyasaDdpToplam = 0
            for item in data_list['icPiyasa']:
                sayfa.cell(satir,column=1,value=item['month']).border = border
                sayfa.cell(satir,column=2,value=item['fob']).border = border
                sayfa.cell(satir,column=3,value=item['ddp']).border = border
                icPiyasaFobToplam += item['fob']
                icPiyasaDdpToplam += item['ddp']
 
                satir += 1
                
            sayfa.cell(satir,column=1,value='Toplam').border = border
            sayfa.cell(satir,column=2,value=icPiyasaFobToplam).border = border
            sayfa.cell(satir,column=3,value=icPiyasaFobToplam).border = border
            
            mekmerFobToplam = 0
            mekmerDdpToplam = 0
            for item in data_list['mekmer']:
                sayfa.cell(satir2,column=5,value=item['month']).border = border
                sayfa.cell(satir2,column=6,value=item['fob']).border = border
                sayfa.cell(satir2,column=7,value=item['ddp']).border = border
                mekmerFobToplam += item['fob']
                mekmerDdpToplam += item['ddp']
 
                satir2 += 1
            
            sayfa.cell(satir2,column=5,value='Toplam').border = border
            sayfa.cell(satir2,column=6,value=mekmerFobToplam).border = border
            sayfa.cell(satir2,column=7,value=mekmerDdpToplam).border = border
            
            kitap.save(target_path)
            kitap.close()
            

            return True

        except Exception as e:
            
            print('get_month_marketing_excel_cikti ExcellÇıktı Hata : ',str(e))
            return False
    
    def get_month_marketing_ayrinti_excel_cikti(self,data_list):
        
        try:
            source_path = 'resource_api/raporlar/sablonlar/monthMarketingAyrintiExcel.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/monthMarketingAyrintiExcel.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            border = Border(left=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
            satir = 2
            icPiyasaFobToplam = 0
            icPiyasaNavlunToplam = 0
            icPiyasaDetay1Toplam = 0
            icPiyasaDetay2Toplam = 0
            icPiyasaDetay3Toplam = 0
            icPiyasaDetay4Toplam = 0
            icPiyasaDdpToplam = 0
            for item in data_list:
                sayfa.cell(satir,column=1,value=item['siparisNo']).border = border
                sayfa.cell(satir,column=2,value=item['fob']).border = border
                sayfa.cell(satir,column=3,value=item['navlun']).border = border
                sayfa.cell(satir,column=4,value=item['detay1']).border = border
                sayfa.cell(satir,column=5,value=item['detay2']).border = border
                sayfa.cell(satir,column=6,value=item['detay3']).border = border
                sayfa.cell(satir,column=7,value=item['detay4']).border = border
                sayfa.cell(satir,column=8,value=item['ddp']).border = border
                icPiyasaFobToplam += item['fob']
                icPiyasaDdpToplam += item['ddp']
                icPiyasaNavlunToplam += item['navlun']
                icPiyasaDetay1Toplam += item['detay1']
                icPiyasaDetay2Toplam += item['detay2']
                icPiyasaDetay3Toplam += item['detay3']
                icPiyasaDetay4Toplam += item['detay4']
 
                satir += 1
                
            sayfa.cell(satir,column=1,value='Toplam').border = border
            sayfa.cell(satir,column=2,value=icPiyasaFobToplam).border = border
            sayfa.cell(satir,column=3,value=icPiyasaNavlunToplam).border = border
            sayfa.cell(satir,column=4,value=icPiyasaDetay1Toplam).border = border
            sayfa.cell(satir,column=5,value=icPiyasaDetay2Toplam).border = border
            sayfa.cell(satir,column=6,value=icPiyasaDetay3Toplam).border = border
            sayfa.cell(satir,column=7,value=icPiyasaDetay4Toplam).border = border
            sayfa.cell(satir,column=8,value=icPiyasaDdpToplam).border = border
            
            
            kitap.save(target_path)
            kitap.close()
            

            return True

        except Exception as e:
            
            print('get_month_marketing_excel_cikti ExcellÇıktı Hata : ',str(e))
            return False
    