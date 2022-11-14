from helpers import SqlConnect,TarihIslemler
from models.raporlar import *
from marshmallow import Schema,fields
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil

class MekusMasraflar():
    def __init__(self):
        self.data = SqlConnect().data
    def getMekusMasraflar(self,year):
        try:
            result = self.data.getStoreList("""
                                                select 


                                                    sum(s.DetayTutar_4) as Mekus,
                                                    s.SiparisNo as SiparisNo


                                                from SiparislerTB s
                                                where s.depo_yukleme=1 and YEAR(s.YuklemeTarihi)=?
                                                group by

                                                    s.MusteriID,s.SiparisNo
                                                order by Mekus desc
                                            """,year)

            liste = list()
            for item in result:
                model = MekusMasraflarModel()
                model.siparisNo = item.SiparisNo
                model.mekusMasraf = item.Mekus
                liste.append(model)
                
            schema = MekusMasraflarSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getMekusMasraflar Hata',str(e))
            
    def getMekusMasraflarExcel(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/mekusMasraflar.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/mekusMasraflar.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')    
            satir = 2

            for item in data:

                sayfa.cell(satir,column=1,value=item['siparisNo'])
                sayfa.cell(satir,column=2,value=item['mekusMasraf'])
            
                satir += 1
                
        
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('getDigerMasraflarExcel  Hata : ',str(e))
            return False
            
class MekusMasraflarSchema(Schema):
    siparisNo = fields.String()
    mekusMasraf = fields.Float()
    
class MekusMasraflarModel:
    siparisNo = ""
    mekusMasraf = 0

    