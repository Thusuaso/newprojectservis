from helpers import SqlConnect,TarihIslemler
from models.raporlar import *
from marshmallow import Schema,fields
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil

class FobMasraflari():
    def __init__(self):
        self.data = SqlConnect().data
    def getFobMasraflari(self,year):
        try:
            result = self.data.getStoreList("""
                                                select 

                                                    sum(sf.Tutar) as Total,
                                                    sf.YuklemeEvrakID as YuklemeEvrakID,
                                                    sf.SiparisFaturaTurID as SiparisFaturaTurID



                                                from SiparisFaturaKayitTB sf

                                                where
                                                    YEAR(sf.Tarih) = ? and sf.Tutar != 0

                                                group by
                                                    sf.YuklemeEvrakID,sf.SiparisFaturaTurID
                                                order by
                                                    Total desc
                                            """,year)
            liste = list()
            for item in result:
                model = FobMaliyetModel()
                if item.YuklemeEvrakID == 70:
                    model.evrakID = item.YuklemeEvrakID
                    model.evrakAdi = 'Gümrük'
                    model.totalFob = item.Total
                elif item.SiparisFaturaTurID ==73:
                    model.evrakID = item.SiparisFaturaTurID
                    model.evrakAdi = 'İlaçlama'
                    model.totalFob = item.Total
                elif item.SiparisFaturaTurID ==15:
                    model.evrakID = item.SiparisFaturaTurID
                    model.evrakAdi = 'Sigorta'
                    model.totalFob = item.Total
                elif item.SiparisFaturaTurID ==9 and item.YuklemeEvrakID == 50:
                    model.evrakID = item.SiparisFaturaTurID
                    model.evrakAdi = 'Liman'
                    model.totalFob = item.Total
                else: 
                    continue
                liste.append(model)
            schema = FobMaliyetSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getFobMasraflari Hata',str(e))
            
    def getFobMasrafExcel(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/fobMasraflar.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/fobMasraflar.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')    
            satir = 2

            for item in data:

                sayfa.cell(satir,column=1,value=item['evrakAdi'])
                sayfa.cell(satir,column=2,value=item['totalFob'])
                satir += 1
                
        
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('getNakliyeExcel  Hata : ',str(e))
            return False
            
class FobMaliyetSchema(Schema):
    evrakID = fields.Int()
    evrakAdi = fields.String()
    totalFob = fields.Float()
    
class FobMaliyetModel:
    evrakID = 0
    evrakAdi = ""
    totalFob = 0
    