
import datetime
from helpers import SqlConnect
from models.finans import MarketingSchema,MarketingModel,BdDepoSchema,BdDepoModel,AylikYuklemeSchema,AylikYuklemeModel,MarketingAyrintiSchema,MarketingAyrintiModel
from openpyxl import *
from openpyxl.styles import Font

import shutil
class Marketing:

    def __init__(self):
        self.data = SqlConnect().data
        self.navlunYukleme = []
        self.navlunUretim = []
        self.liste = []
        self.marketingNavlun = []
    def getMarketingYuklemeHepsi(self):
        result = self.data.getList("""
                                            select 	
                                                sum(su.SatisToplam) as Toplam,
                                                m.Marketing as Marketing
                                            from MusterilerTB m	
                                                inner join SiparislerTB s on s.MusteriID = m.ID
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                            where 	
                                                YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3  and m.Marketing != 'Seleksiyon'
                                            group by
                                                m.Marketing

                                        """)
        
        self.navlunYukleme = self.data.getList("""
                                                select 	
                                                    sum(s.NavlunSatis) as Navlun,
                                                    sum(s.DetayTutar_1) as DetayTutar1,
                                                    sum(s.DetayTutar_2) as DetayTutar2,
                                                    sum(s.DetayTutar_3) as DetayTutar3,
                                                    sum(s.DetayTutar_4) as DetayTutar4,
                                                    
                                                    m.Marketing as Marketing
                                                from MusterilerTB m	
                                                    inner join SiparislerTB s on s.MusteriID = m.ID
                                                where 	
                                                    YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 
                                                group by
                                                    m.Marketing
                                                                                            
                                             
                                             """)
        
        
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = item.Toplam
            model.cfrToplam = item.Toplam + self.__getNavlun(item.Marketing)
            self.liste.append(model)
        # for item in self.__getMarketingYuklemeGhanaHepsi():
        #     self.liste.append(item)
        # for item in self.__getMayaContainerHepsi():
        #     self.liste.append(item)
        
        # for item in self.__getVilloContainerHepsi():
        #     self.liste.append(item)

            
        schema = MarketingSchema(many=True)
        return schema.dump(self.liste)
    
    def getMarketingYukleme(self,month):
        
        result = self.data.getStoreList("""
                                            select 	
                                                sum(su.SatisToplam) as Toplam,
                                                m.Marketing as Marketing
                                            from MusterilerTB m	
                                                inner join SiparislerTB s on s.MusteriID = m.ID
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                            where 	
                                                YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and MONTH(s.YuklemeTarihi) =?
                                            group by
                                                m.Marketing

                                        """,(month))
        
        self.navlunYukleme = self.data.getStoreList("""
                                                select 	
                                                    sum(s.NavlunSatis) as Navlun,
                                                    sum(s.DetayTutar_1) as DetayTutar1,
                                                    sum(s.DetayTutar_2) as DetayTutar2,
                                                    sum(s.DetayTutar_3) as DetayTutar3,
                                                    sum(s.DetayTutar_4) as DetayTutar4,
                                                    
                                                    m.Marketing as Marketing
                                                from MusterilerTB m	
                                                    inner join SiparislerTB s on s.MusteriID = m.ID
                                                where 	
                                                    YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and MONTH(s.YuklemeTarihi) =?
                                                group by
                                                    m.Marketing
                                                                                            
                                             
                                             """,(month))
        
        
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = item.Toplam
            model.cfrToplam = item.Toplam + self.__getNavlun(item.Marketing)
            self.liste.append(model)
        # for item in self.__getMarketingYuklemeGhana(month):
        #     self.liste.append(item)
        # for item in self.__getMayaContainer(month):
        #     self.liste.append(item)
        
        # for item in self.__getVilloContainer(month):
        #     self.liste.append(item)

            
        schema = MarketingSchema(many=True)
        return schema.dump(self.liste)
    
    def getMarketingDetail(self,month):
        try:
            result = self.data.getStoreList("""
                                                select            
                                                    m.ID as MusteriId,            
                                                    m.FirmaAdi as MusteriAdi,               
                                                    m.Marketing,
                                                    (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
                                            
                                    
                                                            
                                        (          
                                        Select Sum(u.SatisToplam) from SiparislerTB s, SiparisUrunTB u where s.SiparisNo=u.SiparisNo and s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) and MONTH(s.YuklemeTarihi) = ?  
                                            
                                        ) as Toplam 


                                                    from            
                                                    MusterilerTB m
                                            
                                            """,(month))
            self.marketingNavlun = self.data.getStoreList("""
                                                select            
                 m.ID as MusteriId,            
                 m.FirmaAdi as MusteriAdi,               
                 m.Marketing,
                (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
          
 
                           
    (          
     Select Sum(s.NavlunSatis) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) and MONTH(s.YuklemeTarihi) = ?  
          
    ) +
	(          
     Select Sum(s.DetayTutar_1) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) and MONTH(s.YuklemeTarihi) = ?
          
    ) +
	(          
     Select Sum(s.DetayTutar_2) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) and MONTH(s.YuklemeTarihi) = ?  
          
    )+
	(          
     Select Sum(s.DetayTutar_3) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) and MONTH(s.YuklemeTarihi) = ?  
          
    ) +
	(          
     Select Sum(s.DetayTutar_4) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) and MONTH(s.YuklemeTarihi) = ?  
          
    ) as Masraflar


                from            
                MusterilerTB m
                                            
                                            """,(month,month,month,month,month))    
        
            liste = list()
            for item in result:
                model = MarketingAyrintiModel()
                if(item.Toplam == None):
                    continue
                else:
                    model.musteri = item.MusteriAdi
                    model.marketing = item.Marketing
                    model.toplamFob = item.Toplam
                    model.toplamCfr = float(item.Toplam) + float(self.__getMarketingDetailNavlun(item.MusteriId))
                    liste.append(model)
            schema = MarketingAyrintiSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getMarketingDetail hata',str(e))
    
    def __getMarketingDetailNavlun(self,musteriId):
        for item in self.marketingNavlun:
            if item.MusteriId != musteriId:
                continue
            else:
                if item.Masraflar != None:
                    return item.Masraflar
                else:
                    return 0
    
    
    def getMarketingDetailHepsi(self):
        try:
            result = self.data.getList("""
                                                select            
                                                    m.ID as MusteriId,            
                                                    m.FirmaAdi as MusteriAdi,               
                                                    m.Marketing,
                                                    (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
                                            
                                    
                                                            
                                        (          
                                        Select Sum(u.SatisToplam) from SiparislerTB s, SiparisUrunTB u where s.SiparisNo=u.SiparisNo and s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) 
                                            
                                        ) as Toplam 


                                                    from            
                                                    MusterilerTB m
                                            
                                            """)
            self.marketingNavlun = self.data.getList("""
                                                select            
                 m.ID as MusteriId,            
                 m.FirmaAdi as MusteriAdi,               
                 m.Marketing,
                (select yu.UlkeAdi from YeniTeklif_UlkeTB yu where yu.Id = m.UlkeId ) as Ulke,
          
 
                           
    (          
     Select Sum(s.NavlunSatis) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) 
          
    ) +
	(          
     Select Sum(s.DetayTutar_1) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate())
          
    ) +
	(          
     Select Sum(s.DetayTutar_2) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate())  
          
    )+
	(          
     Select Sum(s.DetayTutar_3) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate())   
          
    ) +
	(          
     Select Sum(s.DetayTutar_4) from SiparislerTB s where s.SiparisDurumID=3 and s.MusteriID=m.ID and Year(s.YuklemeTarihi)=Year(GetDate()) 
          
    ) as Masraflar


                from            
                MusterilerTB m
                                            
                                            """)    
        
            liste = list()
            for item in result:
                model = MarketingAyrintiModel()
                if(item.Toplam == None):
                    continue
                else:
                    model.musteri = item.MusteriAdi
                    model.marketing = item.Marketing
                    model.toplamFob = item.Toplam
                    model.toplamCfr = float(item.Toplam) + float(self.__getMarketingDetailNavlunHepsi(item.MusteriId))
                    liste.append(model)
            schema = MarketingAyrintiSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getMarketingDetail hata',str(e))
    
    def __getMarketingDetailNavlunHepsi(self,musteriId):
        for item in self.marketingNavlun:
            if item.MusteriId != musteriId:
                continue
            else:
                if item.Masraflar != None:
                    return item.Masraflar
                else:
                    return 0
    
    
    
    
    
    # def __getMarketingYuklemeGhana(self,month):
        
    #     result = self.data.getStoreList("""
    #                                         select 	
    #                                         sum(su.SatisToplam) as Toplam
    #                                         from MusterilerTB m	
    #                                         inner join SiparislerTB s on s.MusteriID = m.ID
    #                                         inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                         YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and MONTH(s.YuklemeTarihi) =? and m.ID=37
    #                                         group by
    #                                         m.UlkeId

    #                                     """,(month))
        
    #     navlunYukleme = self.data.getStoreList("""
    #                                             select 	
    #                                                 sum(s.NavlunSatis) + sum(s.DetayTutar_1) + sum(s.DetayTutar_2) + sum(s.DetayTutar_3) + sum(s.DetayTutar_4) as Navlun

    #                                             from MusterilerTB m	
    #                                                 inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             where 	
    #                                                 YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and MONTH(s.YuklemeTarihi) =? and m.ID=37
    #                                             group by
    #                                                 m.UlkeId
                                                                                            
                                             
    #                                          """,(month))
        
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Ghana'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunYukleme[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste)
    
    
    # def __getMarketingYuklemeGhanaHepsi(self):
        
    #     result = self.data.getList("""
    #                                         select 	
    #                                         sum(su.SatisToplam) as Toplam
    #                                         from MusterilerTB m	
    #                                         inner join SiparislerTB s on s.MusteriID = m.ID
    #                                         inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                         YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and m.ID=37
    #                                         group by
    #                                         m.UlkeId

    #                                     """)
        
    #     navlunYukleme = self.data.getList("""
    #                                             select 	
    #                                                 sum(s.NavlunSatis) + sum(s.DetayTutar_1) + sum(s.DetayTutar_2) + sum(s.DetayTutar_3) + sum(s.DetayTutar_4) as Navlun

    #                                             from MusterilerTB m	
    #                                                 inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             where 	
    #                                                 YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and m.ID=37
    #                                             group by
    #                                                 m.UlkeId
                                                                                            
                                             
    #                                          """)
        
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Ghana'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunYukleme[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste)
    
    
    
    
    def getMarketingUretim(self):
        
        result = self.data.getList("""
                                            select 	
                                                sum(su.SatisToplam) as Toplam,
                                                m.Marketing as Marketing
                                            from MusterilerTB m	
                                                inner join SiparislerTB s on s.MusteriID = m.ID
                                                inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
                                            where 	
                                                    s.SiparisDurumID =2 
                                            group by
                                                m.Marketing

                                        """)
        
        self.navlunUretim = self.data.getList("""
                                                select 	
                                                    sum(s.NavlunSatis) + sum(s.DetayTutar_1) + sum(s.DetayTutar_2) + sum(s.DetayTutar_3) + sum(s.DetayTutar_4) as Navlun,
                                                    m.Marketing as Marketing
                                                from MusterilerTB m	
                                                    inner join SiparislerTB s on s.MusteriID = m.ID
                                                where 	
                                                    s.SiparisDurumID=2
                                                group by
                                                    m.Marketing
                                                                                            
                                             
                                             """)
        
        
        liste = list()
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = item.Toplam
            model.cfrToplam = item.Toplam + self.__getNavlunUretim(item.Marketing)
            liste.append(model)
        # for item in self.__getMarketingUretimGhana():
        #     liste.append(item)
        # for item in self.__getMayaContainerUretim():
        #     liste.append(item)
        # for item in self.__getVilloContainerYukleme():
        #     liste.append(item)
        schema = MarketingSchema(many=True)
        return schema.dump(liste)

    # def __getMarketingUretimGhana(self):
        
    #     result = self.data.getList("""
    #                                         select 	
    #                                         sum(su.SatisToplam) as Toplam
    #                                         from MusterilerTB m	
    #                                         inner join SiparislerTB s on s.MusteriID = m.ID
    #                                         inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                         s.SiparisDurumID =2 and m.ID=37
    #                                         group by
    #                                         m.UlkeId

    #                                     """)
        
    #     navlunYukleme = self.data.getList("""
    #                                             select 	
    #                                                 sum(s.NavlunSatis) + sum(s.DetayTutar_1) + sum(s.DetayTutar_2) + sum(s.DetayTutar_3) + sum(s.DetayTutar_4) as Navlun
    #                                             from MusterilerTB m	
    #                                                 inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             where 	
    #                                                 s.SiparisDurumID =2 and m.ID=37
    #                                             group by
    #                                                 m.UlkeId
                                                                                            
                                             
    #                                          """)
        
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Ghana'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunYukleme[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste)
    

    def getMarketingDepo(self,month):
        result = self.data.getStoreList("""
                                            select 


                                                sum(yu.Total) as Toplam,
                                                yd.CustomersId,
                                                (select ym.CustomersName from YeniDepoMusterilerTB ym where ym.Id = yd.CustomersId) as Marketing


                                            from 
                                                YeniDepoSatisTB yd
                                                inner join YeniDepoSatisUrunlerTB yu on yu.OrderNo = yd.OrderNo

                                            where 
                                                YEAR(yd.Date) = YEAR(GETDATE()) and MONTH(yd.Date) = ? and Payment=1 and Shipped=1

                                            group by
                                                yd.CustomersId
                                        """,(month))
        
        liste = list()
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = float(item.Toplam)  * 0.70
            model.cfrToplam = item.Toplam
            
            liste.append(model)
        schema = MarketingSchema(many=True)
        return schema.dump(liste)
    
    def getMarketingDepoHepsi(self):
        result = self.data.getList("""
                                            select 


                                                sum(yu.Total) as Toplam,
                                                yd.CustomersId,
                                                (select ym.CustomersName from YeniDepoMusterilerTB ym where ym.Id = yd.CustomersId) as Marketing


                                            from 
                                                YeniDepoSatisTB yd
                                                inner join YeniDepoSatisUrunlerTB yu on yu.OrderNo = yd.OrderNo

                                            where 
                                                YEAR(yd.Date) = YEAR(GETDATE()) and Payment=1 and Shipped=1

                                            group by
                                                yd.CustomersId
                                        """)
        
        liste = list()
        for item in result:
            model = MarketingModel()
            model.marketing = item.Marketing
            model.fobToplam = float(item.Toplam)  * 0.70
            model.cfrToplam = item.Toplam
            
            liste.append(model)
        schema = MarketingSchema(many=True)
        return schema.dump(liste)
       
    
    # def __getMayaContainer(self,month):
    #     result = self.data.getStoreList("""
    #                                         select 	
    #                                             sum(su.SatisToplam) as Toplam,
    #                                             m.Marketing as Marketing
    #                                         from MusterilerTB m	
    #                                             inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                             YEAR(s.YuklemeTarihi) = YEAR(GetDate()) and s.SiparisDurumID =3 and MONTH(s.YuklemeTarihi) =? and m.ID in (1373)

    #                                         group by
    #                                             m.Marketing
                                        
    #                                     """,(month))
    #     navlunMaya = self.data.getStoreList("""
    #                                                 select 	
    #                                                     sum(s.NavlunSatis) as Navlun,
    #                                                     m.Marketing as Marketing
    #                                                 from MusterilerTB m	
    #                                                     inner join SiparislerTB s on s.MusteriID = m.ID
    #                                                 where 	
    #                                                     YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and MONTH(s.YuklemeTarihi) =?  and m.ID in (1373)
    #                                                 group by
    #                                                     m.Marketing
                                                 
    #                                              """,(month))
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Maya'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunMaya[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste) 

    # def __getMayaContainerHepsi(self):
    #     result = self.data.getList("""
    #                                         select 	
    #                                             sum(su.SatisToplam) as Toplam,
    #                                             m.Marketing as Marketing
    #                                         from MusterilerTB m	
    #                                             inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                             YEAR(s.YuklemeTarihi) = YEAR(GetDate()) and s.SiparisDurumID =3 and m.ID in (1373)

    #                                         group by
    #                                             m.Marketing
                                        
    #                                     """)
    #     navlunMaya = self.data.getList("""
    #                                                 select 	
    #                                                     sum(s.NavlunSatis) as Navlun,
    #                                                     m.Marketing as Marketing
    #                                                 from MusterilerTB m	
    #                                                     inner join SiparislerTB s on s.MusteriID = m.ID
    #                                                 where 	
    #                                                     YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3  and m.ID in (1373)
    #                                                 group by
    #                                                     m.Marketing
                                                 
    #                                              """)
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Maya'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunMaya[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste) 

    
    # def __getVilloContainer(self,month):
    #     result = self.data.getStoreList("""
    #                                         select 	
    #                                             sum(su.SatisToplam) as Toplam,
    #                                             m.Marketing as Marketing
    #                                         from MusterilerTB m	
    #                                             inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                             YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and MONTH(s.YuklemeTarihi) =? and m.ID in (2400)

    #                                         group by
    #                                             m.Marketing
                                        
    #                                     """,(month))
    #     navlunVillo = self.data.getStoreList("""
    #                                                 select 	
    #                                                     sum(s.NavlunSatis) as Navlun,
    #                                                     m.Marketing as Marketing
    #                                                 from MusterilerTB m	
    #                                                     inner join SiparislerTB s on s.MusteriID = m.ID
    #                                                 where 	
    #                                                     YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and MONTH(s.YuklemeTarihi) =?  and m.ID in (2400)
    #                                                 group by
    #                                                     m.Marketing
                                                 
    #                                              """,(month))
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Villo Home'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunVillo[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste) 
    
    # def __getVilloContainerHepsi(self):
    #     result = self.data.getList("""
    #                                         select 	
    #                                             sum(su.SatisToplam) as Toplam,
    #                                             m.Marketing as Marketing
    #                                         from MusterilerTB m	
    #                                             inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                             YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3 and m.ID in (2400)

    #                                         group by
    #                                             m.Marketing
                                        
    #                                     """)
    #     navlunVillo = self.data.getList("""
    #                                                 select 	
    #                                                     sum(s.NavlunSatis) as Navlun,
    #                                                     m.Marketing as Marketing
    #                                                 from MusterilerTB m	
    #                                                     inner join SiparislerTB s on s.MusteriID = m.ID
    #                                                 where 	
    #                                                     YEAR(s.YuklemeTarihi) = YEAR(GETDATE()) and s.SiparisDurumID =3  and m.ID in (2400)
    #                                                 group by
    #                                                     m.Marketing
                                                 
    #                                              """)
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Villo Home'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunVillo[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste) 
    
    
    # def __getMayaContainerUretim(self):
    #     result = self.data.getList("""
    #                                         select 	
    #                                             sum(su.SatisToplam) as Toplam,
    #                                             m.Marketing as Marketing
    #                                         from MusterilerTB m	
    #                                             inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                              s.SiparisDurumID =2 and m.ID in (1373)

    #                                         group by
    #                                             m.Marketing
                                        
    #                                     """)
    #     navlunMaya = self.data.getList("""
    #                                                 select 	
    #                                                     sum(s.NavlunSatis) as Navlun,
    #                                                     m.Marketing as Marketing
    #                                                 from MusterilerTB m	
    #                                                     inner join SiparislerTB s on s.MusteriID = m.ID
    #                                                 where 	
    #                                                     s.SiparisDurumID =2   and m.ID in (1373)
    #                                                 group by
    #                                                     m.Marketing
                                                 
    #                                              """)
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Maya'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunMaya[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste) 

    
    # def __getVilloContainerYukleme(self):
    #     result = self.data.getList("""
    #                                         select 	
    #                                             sum(su.SatisToplam) as Toplam,
    #                                             m.Marketing as Marketing
    #                                         from MusterilerTB m	
    #                                             inner join SiparislerTB s on s.MusteriID = m.ID
    #                                             inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo
    #                                         where 	
    #                                             s.SiparisDurumID =2  and m.ID in (2400)

    #                                         group by
    #                                             m.Marketing
                                        
    #                                     """)
    #     navlunVillo = self.data.getList("""
    #                                                 select 	
    #                                                     sum(s.NavlunSatis) as Navlun,
    #                                                     m.Marketing as Marketing
    #                                                 from MusterilerTB m	
    #                                                     inner join SiparislerTB s on s.MusteriID = m.ID
    #                                                 where 	
    #                                                     s.SiparisDurumID =2  and m.ID in (2400)
    #                                                 group by
    #                                                     m.Marketing
                                                 
    #                                              """)
        
    #     liste = list()
    #     for item in result:
    #         model = MarketingModel()
    #         model.marketing = 'Villo Home'
    #         model.fobToplam = item.Toplam
    #         model.cfrToplam = item.Toplam + navlunVillo[0].Navlun
    #         liste.append(model)
    #     schema = MarketingSchema(many=True)
    #     return schema.dump(liste) 
    
    
    def __getNavlun(self,marketing):
        for item in self.navlunYukleme:
            if item.Marketing != marketing:
                continue
            else:
                return item.Navlun + item.DetayTutar1 + item.DetayTutar2 + item.DetayTutar3 + item.DetayTutar4
            
            
    def __getNavlunUretim(self,marketing):
        for item in self.navlunUretim:
            if item.Marketing != marketing:
                continue
            else:
                return item.Navlun
            
            
    def getBdDepoList(self):
        result =self.data.getList("""
                            select 


                                sum(ysu.Total) as Toplam,
                                MONTH(ys.Date) as Ay



                            from YeniDepoSatisTB ys
                            inner join YeniDepoSatisUrunlerTB ysu on ys.OrderNo = ysu.OrderNo
                            where YEAR(ys.Date) = YEAR(GETDATE())  and ys.Shipped=1 and ys.CustomersId=1
                            group by MONTH(ys.Date)
                            order by MONTH(ys.Date)
                            
                          
                          
                          
                          """)
        liste = list()
        for item in result:
            model = BdDepoModel()
            model.ay = self.__getMonth(item.Ay)
            model.cfrToplam = item.Toplam
            model.fobToplam = round(float(item.Toplam) * 0.7, 2)
            liste.append(model)
        
        schema = BdDepoSchema(many=True)
        return schema.dump(liste)
    
    
    def getYuklemeAylikList(self):
        
        date = datetime.datetime.now()
        month = date.month + 1
        liste = list()
        
        for item in range(1,month):
            result = self.data.getStoreList("""
                                    select  
                                    s.YuklemeTarihi,  
                                    s.SiparisNo,  
                                    m.FirmaAdi as MusteriAdi,  
                                    (select Sum(SatisToplam) from SiparisUrunTB su where su.SiparisNo=s.SiparisNo) as Fob,  
                                    (select Sum(SatisToplam) from SiparisUrunTB su where su.SiparisNo=s.SiparisNo)+  
                                    dbo.Get_SiparisNavlun(s.SiparisNo) as Dtp,  
                                    'Konteyner' as Tur,m.Marketing  
                                    from  
                                    SiparislerTB s,MusterilerTB m  
                                    where Year(YuklemeTarihi)=YEAR(GETDATE())
                                    and Month(YuklemeTarihi)=?
                                    and m.ID=s.MusteriID  
                                    and m.Marketing not in ('Mekmar Numune','Seleksiyon','Warehouse')  
                                    and m.Marketing is not null  
                                    
                                    union  
                                    select  
                                    s.Tarih as YuklemeTarihi,  
                                    s.CikisNo as SiparisNo,  
                                    m.FirmaAdi as MusteriAdi,  
                                    Sum(Toplam) as Fob  
                                    ,Sum((s.BirimFiyat+7.5)*u.Miktar) as Dtp,  
                                    'Depo' as Tur,m.Marketing  
                                    from  
                                    SevkiyatTB s,MusterilerTB m,UretimTB u  
                                    where s.MusteriID=m.ID and u.KasaNo=s.KasaNo  
                                    and Year(s.Tarih)=YEAR(GETDATE()) and Month(s.Tarih)=?
                                    and m.Mt_No=1  
                                    group by  
                                    s.Tarih,s.CikisNo,m.FirmaAdi,m.Marketing 
                                   
                                   """,(item,item))
            model = AylikYuklemeModel() 
            for item2 in result:
                model.ay = self.__getMonth(item)
                model.fobToplam += item2.Fob
                model.cfrToplam += item2.Dtp
            liste.append(model)
            
        schema = AylikYuklemeSchema(many=True)
        return schema.dump(liste)
            
    def __getMonth(self,month):
            monthName = ['', 'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran', 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
            return monthName[month]

    def byMarketingExcellCikti(self,data):
        try:
        
            source_path = r"resource_api/raporlar/sablonlar/by_marketing_excel.xlsx"
            target_path = r"resource_api/raporlar/dosyalar/by_marketing_excel.xlsx"

            shutil.copy2(source_path,target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 3
            satir2 = 3
            satir3 = 3
            toplamMekmarFob = 0
            toplamMekmarCfr = 0
            for item in data['mekmarLoadMonths']:

                sayfa.cell(satir,column=1,value=item['ay'])
                sayfa.cell(satir,column=2,value=item['fobToplam'])
                sayfa.cell(satir,column=3,value=item['cfrToplam'])
                toplamMekmarFob += item['fobToplam']
                toplamMekmarCfr += item['cfrToplam']
                
                satir += 1
            toplam = sayfa.cell(satir,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=2,value=toplamMekmarFob).font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=3,value=toplamMekmarCfr).font = Font(color='f44336',bold=True)
            
            
            toplamFobBd= 0
            toplamCfrBd = 0
            for item in data['mekmarBdLoadMonths']:
                sayfa.cell(satir2,column=5,value=item['ay'])
                sayfa.cell(satir2,column=6,value=item['fobToplam'])
                sayfa.cell(satir2,column=7,value=item['cfrToplam'])
                toplamFobBd += item['fobToplam']
                toplamCfrBd += item['cfrToplam']
                satir2 += 1
                
            sayfa.cell(satir2,column=5,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa.cell(satir2,column=6,value=toplamFobBd).font = Font(color='f44336',bold=True)
            sayfa.cell(satir2,column=7,value=toplamCfrBd).font = Font(color='f44336',bold=True)
                
            toplamFobTotal = 0
            toplamCfrTotal = 0
            for item in data['mekmarTotalLoadMonths']:
                sayfa.cell(satir3,column=9,value=item['ay'])
                sayfa.cell(satir3,column=10,value=item['fobToplami'])
                sayfa.cell(satir3,column=11,value=item['cfrToplami'])
                toplamFobTotal += item['fobToplami']
                toplamCfrTotal += item['cfrToplami']
                
                satir3 += 1
                
            sayfa.cell(satir3,column=9,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa.cell(satir3,column=10,value=toplamFobTotal).font = Font(color='f44336',bold=True)
            sayfa.cell(satir3,column=11,value=toplamCfrTotal).font = Font(color='f44336',bold=True)
            
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem byMarketingExcellCikti Hata : ',str(e))
            return False
    
    def byCustomersExcellCikti(self,data):
        try:
            source_path = r"resource_api/raporlar/sablonlar/by_customers_excel.xlsx"
            target_path = r"resource_api/raporlar/dosyalar/by_customers_excel.xlsx"

            shutil.copy2(source_path,target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            sayfa2 = kitap.get_sheet_by_name('Sayfa2')

            satir = 2
  
            toplamSatisBuyil = 0
            toplamSatisGecenyil = 0
            toplamSatisOncekiyil = 0
            toplamSatisOnDokuzyil = 0
            toplamSatisTotal = 0
            toplamSatisTotalCfr = 0
            
            for item in data['byCustomersProduct']:

                sayfa.cell(satir,column=1,value=item['musteriAdi'])
                sayfa.cell(satir,column=2,value=item['marketing'])
                sayfa.cell(satir,column=3,value=item['ulkeAdi'])
                
                if(item['satisToplamiBuYil'] != None):
                    sayfa.cell(satir,column=4,value=item['satisToplamiBuYil'])
                else:
                    sayfa.cell(satir,column=4,value=0)
                
                if(item['satisToplamiGecenYil'] != None):
                    sayfa.cell(satir,column=5,value=item['satisToplamiGecenYil'])
                else:
                    sayfa.cell(satir,column=5,value=0)
                
                if(item['satisToplamiOncekiYil'] != None):
                    sayfa.cell(satir,column=6,value=item['satisToplamiOncekiYil'])
                else:
                    sayfa.cell(satir,column=6,value=0)
                    
                if(item['satisToplamiOnDokuzYil'] != None):
                    sayfa.cell(satir,column=7,value=item['satisToplamiOnDokuzYil'])
                else:
                    sayfa.cell(satir,column=7,value=0)
                
                if(item['toplam'] != None):
                    sayfa.cell(satir,column=8,value=item['toplam'])
                else:
                    sayfa.cell(satir,column=8,value=0)
                    
                if(item['toplam'] != None):
                    sayfa.cell(satir,column=9,value=item['toplamCfr'])
                else:
                    sayfa.cell(satir,column=9,value=0)

                if(item['satisToplamiBuYil'] != None):
                    toplamSatisBuyil += item['satisToplamiBuYil']
                if(item['satisToplamiGecenYil'] != None):
                    toplamSatisGecenyil += item['satisToplamiGecenYil']
                if(item['satisToplamiOncekiYil'] != None):
                    toplamSatisOncekiyil += item['satisToplamiOncekiYil']
                if(item['satisToplamiOnDokuzYil'] != None):
                    toplamSatisOnDokuzyil += item['satisToplamiOnDokuzYil']
                if(item['toplam'] != None):
                    toplamSatisTotal += item['toplam']
                    
                if(item['toplamCfr'] != None):
                    toplamSatisTotalCfr += item['toplamCfr']

                

                
                satir += 1
            sayfa.cell(satir,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=4,value=self.__getNoneType(toplamSatisBuyil)).font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=5,value=self.__getNoneType(toplamSatisGecenyil)).font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=6,value=self.__getNoneType(toplamSatisOncekiyil)).font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=7,value=self.__getNoneType(toplamSatisOnDokuzyil)).font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=8,value=self.__getNoneType(toplamSatisTotal)).font = Font(color='f44336',bold=True) 
            sayfa.cell(satir,column=9,value=self.__getNoneType(toplamSatisTotalCfr)).font = Font(color='f44336',bold=True)
            
            satir2 = 3
            buyilToplam = 0
            toplamFob = 0
            toplamCfr = 0
            for item in data['imperialHomes']:
                sayfa2.cell(satir2,column=1,value=item['musteriAdi'])
                sayfa2.cell(satir2,column=2,value=self.__getNoneType(item['satisToplamiBuYil']))
                sayfa2.cell(satir2,column=3,value=self.__getNoneType(item['toplam']))
                sayfa2.cell(satir2,column=4,value=self.__getNoneType(item['toplamCfr']))
                buyilToplam += self.__getNoneType(item['satisToplamiBuYil'])
                toplamFob += self.__getNoneType(item['toplam'])
                toplamCfr += self.__getNoneType(item['toplamCfr'])
                
                satir2 += 1
            sayfa2.cell(satir2,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa2.cell(satir2,column=2,value=buyilToplam).font = Font(color='f44336',bold=True)
            sayfa2.cell(satir2,column=3,value=toplamFob).font = Font(color='f44336',bold=True)
            sayfa2.cell(satir2,column=4,value=toplamCfr).font = Font(color='f44336',bold=True)
            
            
            satir3 = 3
            buyilToplamMekmar = 0
            toplamFobMekmar = 0
            toplamCfrMekmar = 0
            for item in data['mekmar']:
                sayfa2.cell(satir3,column=6,value=item['musteriAdi'])
                sayfa2.cell(satir3,column=7,value=self.__getNoneType(item['satisToplamiBuYil']))
                sayfa2.cell(satir3,column=8,value=self.__getNoneType(item['toplam']))
                sayfa2.cell(satir3,column=9,value=self.__getNoneType(item['toplamCfr']))
                buyilToplamMekmar += self.__getNoneType(item['satisToplamiBuYil'])
                toplamFobMekmar += self.__getNoneType(item['toplam'])
                toplamCfrMekmar += self.__getNoneType(item['toplamCfr'])
                
                satir3 += 1
            sayfa2.cell(satir3,column=6,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa2.cell(satir3,column=7,value=self.__getNoneType(buyilToplamMekmar)).font = Font(color='f44336',bold=True)
            sayfa2.cell(satir3,column=8,value=self.__getNoneType(toplamFobMekmar)).font = Font(color='f44336',bold=True)
            sayfa2.cell(satir3,column=9,value=self.__getNoneType(toplamCfrMekmar)).font = Font(color='f44336',bold=True)
            
            satir4 = 3
            buyilToplamIcPiyasa = 0
            toplamFobIcPiyasa = 0
            toplamCfrIcPiyasa = 0
            for item in data['icPiyasa']:
                sayfa2.cell(satir4,column=11,value=item['musteriAdi'])
                sayfa2.cell(satir4,column=12,value=self.__getNoneType(item['satisToplamiBuYil']))
                sayfa2.cell(satir4,column=13,value=self.__getNoneType(item['toplam']))
                sayfa2.cell(satir4,column=14,value=self.__getNoneType(item['toplamCfr']))
                buyilToplamIcPiyasa += self.__getNoneType(item['satisToplamiBuYil'])
                toplamFobIcPiyasa += self.__getNoneType(item['toplam'])
                toplamCfrIcPiyasa += self.__getNoneType(item['toplamCfr'])
                
                satir4 += 1
            sayfa2.cell(satir4,column=11,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa2.cell(satir4,column=12,value=buyilToplamIcPiyasa).font = Font(color='f44336',bold=True)
            sayfa2.cell(satir4,column=13,value=toplamFobIcPiyasa).font = Font(color='f44336',bold=True)
            sayfa2.cell(satir4,column=14,value=toplamCfrIcPiyasa).font = Font(color='f44336',bold=True)
                

            
            satir5 = 3
            buyilToplamMekmer = 0
            toplamFobMekmer = 0
            toplamCfrMekmer = 0
            for item in data['mekmer']:
                sayfa2.cell(satir5,column=16,value=item['musteriAdi'])
                sayfa2.cell(satir5,column=17,value=self.__getNoneType(item['satisToplamiBuYil']))
                sayfa2.cell(satir5,column=18,value=self.__getNoneType(item['toplam']))
                sayfa2.cell(satir5,column=19,value=self.__getNoneType(item['toplamCfr']))
                buyilToplamMekmer += self.__getNoneType(item['satisToplamiBuYil'])
                toplamFobMekmer += self.__getNoneType(item['toplam'])
                toplamCfrMekmer += self.__getNoneType(item['toplamCfr'])
                
                satir5 += 1
            sayfa2.cell(satir5,column=16,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa2.cell(satir5,column=17,value=buyilToplamMekmer).font = Font(color='f44336',bold=True)
            sayfa2.cell(satir5,column=18,value=toplamFobMekmer).font = Font(color='f44336',bold=True)
            sayfa2.cell(satir5,column=19,value=toplamCfrMekmer).font = Font(color='f44336',bold=True)
                
            
            

            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem byMarketingExcellCikti Hata : ',str(e))
            return False
    
    def byMarketingDetailExcellCikti(self,data):
        try:
            source_path = r"resource_api/raporlar/sablonlar/by_marketing_detail_excel.xlsx"
            target_path = r"resource_api/raporlar/dosyalar/by_marketing_detail_excel.xlsx"

            shutil.copy2(source_path,target_path)

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            sayfa2 = kitap.get_sheet_by_name('Sayfa2')

            satir = 3
            satir2 = 3
            satir3 = 3
            toplamYuklemeFob = 0
            toplamYuklemeCfr = 0
            
            for item in data['byMarketingLoadMonth']:

                sayfa.cell(satir,column=1,value=item['marketing'])
                
                if(item['fobToplam'] != None):
                    sayfa.cell(satir,column=2,value=item['fobToplam'])
                else:
                    sayfa.cell(satir,column=2,value=0)
                
                if(item['cfrToplam'] != None):
                    sayfa.cell(satir,column=3,value=item['cfrToplam'])
                else:
                    sayfa.cell(satir,column=3,value=0)

                toplamYuklemeFob += self.__getNoneType(item['fobToplam'])
                toplamYuklemeCfr += self.__getNoneType(item['cfrToplam'])
                
                
                satir += 1
            sayfa.cell(satir,column=1,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=2,value=toplamYuklemeFob).font = Font(color='f44336',bold=True)
            sayfa.cell(satir,column=3,value=toplamYuklemeCfr).font = Font(color='f44336',bold=True)
            
            
            toplamYuklemeFobWarehouse = 0
            toplamYuklemeCfrWarehouse = 0
            for item in data['byMarketingWarehouseLoad']:

                sayfa.cell(satir3,column=5,value=item['marketing'])
                
                if(item['fobToplam'] != None):
                    sayfa.cell(satir3,column=6,value=item['fobToplam'])
                else:
                    sayfa.cell(satir3,column=6,value=0)
                
                if(item['cfrToplam'] != None):
                    sayfa.cell(satir3,column=7,value=item['cfrToplam'])
                else:
                    sayfa.cell(satir3,column=7,value=0)

                toplamYuklemeFobWarehouse += self.__getNoneType(item['fobToplam'])
                toplamYuklemeCfrWarehouse += self.__getNoneType(item['cfrToplam'])
                
                
                satir3 += 1
            sayfa.cell(satir3,column=5,value='Toplam').font = Font(color='f44336',bold=True)
            sayfa.cell(satir3,column=6,value=toplamYuklemeFobWarehouse).font = Font(color='f44336',bold=True)
            sayfa.cell(satir3,column=7,value=toplamYuklemeCfrWarehouse).font = Font(color='f44336',bold=True)
            
            toplamYuklemeFobImperialHomes = 0
            toplamYuklemeCfrImperialHomes = 0
            satir4 = 3
            if(len(data['imperialHomes'])>0):
                
                for item in data['imperialHomes']:

                    sayfa2.cell(satir4,column=1,value=item['musteri'])
                    
                    if(item['toplamFob'] != None):
                        sayfa2.cell(satir4,column=2,value=item['toplamFob'])
                    else:
                        sayfa2.cell(satir4,column=2,value=0)
                    
                    if(item['toplamCfr'] != None):
                        sayfa2.cell(satir4,column=3,value=item['toplamCfr'])
                    else:
                        sayfa2.cell(satir4,column=3,value=0)

                    toplamYuklemeFobImperialHomes += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrImperialHomes += self.__getNoneType(item['toplamCfr'])
                    
                    
                    satir4 += 1
                sayfa2.cell(satir4,column=1,value='Toplam').font = Font(color='f44336',bold=True)
                sayfa2.cell(satir4,column=2,value=toplamYuklemeCfrImperialHomes).font = Font(color='f44336',bold=True)
                sayfa2.cell(satir4,column=3,value=toplamYuklemeCfrImperialHomes).font = Font(color='f44336',bold=True)
            
            toplamYuklemeFobMekmar = 0
            toplamYuklemeCfrMekmar = 0
            satir5 = 3
            if(len(data['mekmar'])>0):
                
                for item in data['mekmar']:

                    sayfa2.cell(satir5,column=5,value=item['musteri'])
                    
                    if(item['toplamFob'] != None):
                        sayfa2.cell(satir5,column=6,value=item['toplamFob'])
                    else:
                        sayfa2.cell(satir5,column=6,value=0)
                    
                    if(item['toplamCfr'] != None):
                        sayfa2.cell(satir5,column=7,value=item['toplamCfr'])
                    else:
                        sayfa2.cell(satir5,column=7,value=0)

                    toplamYuklemeFobMekmar += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrMekmar += self.__getNoneType(item['toplamCfr'])
                    
                    
                    satir5 += 1
                sayfa2.cell(satir5,column=5,value='Toplam').font = Font(color='f44336',bold=True)
                sayfa2.cell(satir5,column=6,value=toplamYuklemeFobMekmar).font = Font(color='f44336',bold=True)
                sayfa2.cell(satir5,column=7,value=toplamYuklemeCfrMekmar).font = Font(color='f44336',bold=True)
            
            toplamYuklemeFobMekmer = 0
            toplamYuklemeCfrMekmer = 0
            satir6 = 3
            if(len(data['mekmer'])>0):
                
                for item in data['mekmer']:

                    sayfa2.cell(satir6,column=9,value=item['musteri'])
                    
                    if(item['toplamFob'] != None):
                        sayfa2.cell(satir6,column=10,value=item['toplamFob'])
                    else:
                        sayfa2.cell(satir6,column=10,value=0)
                    
                    if(item['toplamCfr'] != None):
                        sayfa2.cell(satir6,column=11,value=item['toplamCfr'])
                    else:
                        sayfa2.cell(satir6,column=11,value=0)

                    toplamYuklemeFobMekmer += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrMekmer += self.__getNoneType(item['toplamCfr'])
                    
                    
                    satir6 += 1
                sayfa2.cell(satir6,column=9,value='Toplam').font = Font(color='f44336',bold=True)
                sayfa2.cell(satir6,column=10,value=toplamYuklemeFobMekmer).font = Font(color='f44336',bold=True)
                sayfa2.cell(satir6,column=11,value=toplamYuklemeCfrMekmer).font = Font(color='f44336',bold=True)
               
            toplamYuklemeFobEfes = 0
            toplamYuklemeCfrEfes = 0
            satir7 = 3
            if(len(data['efes'])>0):
                
                for item in data['efes']:

                    sayfa2.cell(satir7,column=13,value=item['musteri'])
                    
                    if(item['toplamFob'] != None):
                        sayfa2.cell(satir7,column=14,value=item['toplamFob'])
                    else:
                        sayfa2.cell(satir7,column=14,value=0)
                    
                    if(item['toplamCfr'] != None):
                        sayfa2.cell(satir7,column=15,value=item['toplamCfr'])
                    else:
                        sayfa2.cell(satir7,column=15,value=0)

                    toplamYuklemeFobEfes += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrEfes += self.__getNoneType(item['toplamCfr'])
                    
                    
                    satir7 += 1
                sayfa2.cell(satir7,column=13,value='Toplam').font = Font(color='f44336',bold=True)
                sayfa2.cell(satir7,column=14,value=toplamYuklemeFobMekmer).font = Font(color='f44336',bold=True)
                sayfa2.cell(satir7,column=15,value=toplamYuklemeCfrMekmer).font = Font(color='f44336',bold=True)
            
            
            toplamYuklemeFobIcPiyasa = 0
            toplamYuklemeCfrIcPiyasa = 0
            satir8 = 3
            if(len(data['icPiyasa'])>0):
                
                for item in data['icPiyasa']:

                    sayfa2.cell(satir8,column=17,value=item['musteri'])
                    
                    if(item['toplamFob'] != None):
                        sayfa2.cell(satir8,column=18,value=item['toplamFob'])
                    else:
                        sayfa2.cell(satir8,column=18,value=0)
                    
                    if(item['toplamCfr'] != None):
                        sayfa2.cell(satir8,column=19,value=item['toplamCfr'])
                    else:
                        sayfa2.cell(satir8,column=19,value=0)

                    toplamYuklemeFobIcPiyasa += self.__getNoneType(item['toplamFob'])
                    toplamYuklemeCfrIcPiyasa += self.__getNoneType(item['toplamCfr'])
                    
                    
                    satir8 += 1
                sayfa2.cell(satir8,column=17,value='Toplam').font = Font(color='f44336',bold=True)
                sayfa2.cell(satir8,column=18,value=toplamYuklemeFobIcPiyasa).font = Font(color='f44336',bold=True)
                sayfa2.cell(satir8,column=19,value=toplamYuklemeCfrIcPiyasa).font = Font(color='f44336',bold=True)
               
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem byMarketingExcellCikti Hata : ',str(e))
            return False
    
    
    
    
    
    
    def __getNoneType(self,value):
        if value == None:
            return 0
        else:
            return value