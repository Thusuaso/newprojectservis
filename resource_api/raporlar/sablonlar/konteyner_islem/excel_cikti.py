from openpyxl import *
import shutil


class ExcelCiktiIslem:

    def konteynerCikti(self,data_list):

        try:
            source_path = 'resource_api/finans/konteyner_islem/sablonlar/konteyner_listesi.xlsx'
            target_path = 'resource_api/finans/konteyner_islem/dosyalar/konteyner_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')

            satir = 3
            data_list = sorted(data_list, key=lambda x:x['genel_bakiye'],reverse=True)
            for item in data_list:
                
                sayfa.cell(satir,column=1,value=item['musteriadi'])
                sayfa.cell(satir,column=2,value=item['kapanmayan_siparis'])
                sayfa.cell(satir,column=3,value=item['kapanmayan_odenen'])
                sayfa.cell(satir,column=4,value=item['kapanmayan_kalan'])
                


                sayfa.cell(satir,column=5,value=item['pesinat'] + item['eski_pesinat'])
                sayfa.cell(satir,column=6,value=item['genel_bakiye'])

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem konteynerCikti Hata : ',str(e))
            return False

    def depoCikti(self,data_list):

        try:
            source_path = 'resource_api/finans/konteyner_islem/sablonlar/depo_listesi.xlsx'
            target_path = 'resource_api/finans/konteyner_islem/dosyalar/depo_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')

            satir = 3

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['musteriadi'])               
                sayfa.cell(satir,column=2,value=item['ciro'])
                sayfa.cell(satir,column=3,value=item['odenen'])
                sayfa.cell(satir,column=4,value=item['bakiye'])

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False

    def konteyner_ayrinti_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/finans/konteyner_islem/sablonlar/konteyner_ayrinti_listesi.xlsx'
            target_path = 'resource_api/finans/konteyner_islem/dosyalar/konteyner_ayrinti_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')

            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['siparisno'])               
                sayfa.cell(satir,column=2,value=item['yuklemetarihi'])
                sayfa.cell(satir,column=3,value=item['tip'])
                sayfa.cell(satir,column=4,value=item['toplam'])
                sayfa.cell(satir,column=5,value=item['kalan'])
                sayfa.cell(satir,column=6,value=item['vade'])


                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False

    def konteyner_odeme_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/finans/konteyner_islem/sablonlar/konteyner_odeme_listesi.xlsx'
            target_path = 'resource_api/finans/konteyner_islem/dosyalar/konteyner_odeme_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['tarih'])               
                sayfa.cell(satir,column=2,value=item['tutar'])

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False

    def musteri_odeme_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/finans/konteyner_islem/sablonlar/musteri_odeme_listesi.xlsx'
            target_path = 'resource_api/finans/konteyner_islem/dosyalar/musteri_odeme_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['tarih'])   
                sayfa.cell(satir,column=2,value=item['musteriadi'])             
                sayfa.cell(satir,column=3,value=item['tutar'])

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False
    