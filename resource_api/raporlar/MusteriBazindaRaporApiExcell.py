from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil



class MusteriBazindaExcellCikti:

    def musteri_bazinda_rapor_cikti_tekli(self,data_list):
        try:
            source_path = 'resource_api/raporlar/sablonlar/musterisip_bazinda_rapor.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/musterisip_bazinda_rapor.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
            for item in data_list:
                
                sayfa.cell(satir,column=1,value=item['musteri'])  
                sayfa.cell(satir,column=2,value=item['fob'])             
                sayfa.cell(satir,column=3,value=item['ddp'])

                satir += 1
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcellMusteriBazindaCikti musteribazinda Hata : ',str(e))
            return False

    def musteri_bazinda_rapor_cikti_coklu(self,datas,dataSum):
        try:
            source_path = 'resource_api/raporlar/sablonlar/musterisip_bazinda_rapor.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/musterisip_bazinda_rapor.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
            satir2= 2
            for item in datas:
                sayfa.cell(satir,column=1,value=item['musteri'])  
                sayfa.cell(satir,column=2,value=item['fob'])             
                sayfa.cell(satir,column=3,value=item['ddp'])
                

                satir += 1
                
            for item2 in dataSum:
                sayfa.cell(satir2,column=5,value=item2['musteri'])  
                sayfa.cell(satir2,column=6,value=item2['fob'])             
                sayfa.cell(satir2,column=7,value=item2['ddp'])
                

                satir2 += 1
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcellMusteriBazindaCikti musteribazinda Hata : ',str(e))
            return False
    
    def musteri_bazinda_toplam_excell(self,data_list):
        try:
            source_path = 'resource_api/raporlar/sablonlar/musterisip_bazinda_rapor_toplam.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/musterisip_bazinda_rapor_toplam.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
            
            for item in data_list:
                sayfa.cell(satir,column=1,value=item['musteri'])  
                sayfa.cell(satir,column=2,value=item['temsilci'])
                sayfa.cell(satir,column=3,value=item['ulkeAdi'])
                sayfa.cell(satir,column=4,value=item['marketing'])
                sayfa.cell(satir,column=5,value=item['Toplam'])
                sayfa.cell(satir,column=6,value=item['BuYilUretim'])
                sayfa.cell(satir,column=7,value=item['BuYilSevkiyat'])
                sayfa.cell(satir,column=8,value=item['GecenYil'])
                sayfa.cell(satir,column=9,value=item['OncekiYil'])
                sayfa.cell(satir,column=10,value=item['OnDokuzYili'])
                sayfa.cell(satir,column=11,value=item['OnSekizYili'])
                sayfa.cell(satir,column=12,value=item['OnYediYili'])
                sayfa.cell(satir,column=13,value=item['OnAltiYili'])
                sayfa.cell(satir,column=14,value=item['OnBesYili'])
                sayfa.cell(satir,column=15,value=item['OnDortYili'])
                sayfa.cell(satir,column=16,value=item['OnUcYili'])
                sayfa.cell(satir,column=16,value=item['OnUcYiliOncesi'])
                satir += 1
                
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcellMusteriBazindaCikti musteribazinda Hata : ',str(e))
            return False