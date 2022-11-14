from helpers import SqlConnect,TarihIslemler
from models.raporlar import *
from marshmallow import Schema,fields
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil

class BankaVeEvrakMasraflar():
    def __init__(self):
        self.data = SqlConnect().data
    def getBankaVeEvrakMasraflar(self,year):
        try:
            result = self.data.getStoreList("""
                                                select 
                                                    sum(o.Masraf) as bankaMasrafi,
                                                    (select m.FirmaAdi from MusterilerTB m where m.ID=o.MusteriID) as Musteri,
                                                    sum(s.EvrakGideri) as kuryeMasrafi

                                                from OdemelerTB o,SiparislerTB s

                                                where o.MusteriID = s.MusteriID  and YEAR(Tarih)=?

                                                group by
                                                    o.MusteriID
                                                    
                                                order by 
                                                    bankaMasrafi desc
                                            """,year)
            liste = list()
            for item in result:
                model = BankaVeKuryeMaliyetModel()
                if item.bankaMasrafi == 0 and item.kuryeMasrafi == 0:
                    continue
                else:
                    
                    model.musteri = item.Musteri
                    model.banka = item.bankaMasrafi
                    model.evrak = item.kuryeMasrafi
                
                liste.append(model)
            schema = BankaVeKuryeMaliyetSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getBankaVeEvrakMasraflar Hata',str(e))
            
    def getBankaVeEvrakMasraflarExcel(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/bankaVeEvrakMasraflar.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/bankaVeEvrakMasraflar.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')    
            satir = 2
            for item in data:

                sayfa.cell(satir,column=1,value=item['musteri'])
                sayfa.cell(satir,column=2,value=item['banka'])
                sayfa.cell(satir,column=3,value=item['evrak'])
                satir += 1
                
        
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('getBankaVeEvrakMasraflarExcel  Hata : ',str(e))
            return False
            
class BankaVeKuryeMaliyetSchema(Schema):
    musteri = fields.String()
    banka = fields.Float()
    evrak = fields.Float()
    
class BankaVeKuryeMaliyetModel:
    musteri = ""
    banka = 0
    evrak = 0
    