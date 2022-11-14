from helpers import SqlConnect,TarihIslemler
from models.raporlar import *
from marshmallow import Schema,fields
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil
class UretimTedarikci:
    def __init__(self):
        self.data = SqlConnect().data

    def getUretimTedarikci(self,year):
        try:
            result = self.data.getStoreList("""
                                                select 
                                                sum(su.SatisToplam) as Toplam,
                                                (select t.FirmaAdi from TedarikciTB t where t.ID = su.TedarikciID) as Tedarikci


                                                from 

                                                    SiparislerTB s
                                                    inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo

                                                where YEAR(s.YuklemeTarihi) = ?

                                                group by 
                                                    su.TedarikciID
                                                order by
	                                                Toplam desc
                                            """,(year))
            liste = list()
            for item in result:
                model = UretimTedarikciModel()
                model.tedarikci = item.Tedarikci
                model.topUretimBedel = item.Toplam
                liste.append(model)
            schema = UretimTedarikciSchema(many=True)
            return schema.dump(liste)
        except Exception as e:
            print('getUretimTedarikci hata',str(e))
            return False

    def getUretimTedarikciExcel(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/uretimTedarikciRapor.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/uretimTedarikciRapor.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')    
            satir = 2

            for item in data:

                sayfa.cell(satir,column=1,value=item['tedarikci'])
                sayfa.cell(satir,column=2,value=item['topUretimBedel'])
                satir += 1
                
        
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('getUretimTedarikciExcel  Hata : ',str(e))
            return False


    
    
class UretimTedarikciSchema(Schema):
    tedarikci = fields.String()
    topUretimBedel = fields.Float()
    
    
class UretimTedarikciModel():
    tedarikci = ""
    topUretimBedel = 0
