from helpers import SqlConnect,TarihIslemler
from models.raporlar import *
from marshmallow import Schema,fields
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil

class NavlunMasraflar():
    def __init__(self):
        self.data = SqlConnect().data
    def getNavlunMasraflar(self,year):
        try:
            result = self.data.getStoreList("""
                                                select (select a.FirmaAdi from FirmalarTB a where a.ID=k.FirmaID)  as firma,
        
                                                sum(f.Tutar) as TotalNavlun
                                                from SiparisFaturaKayitTB f , KonteynerDigerFaturalarKayitTB k 
                                                where k.ID=f.FaturaKayitID and f.SiparisFaturaTurID !=0 and f.SiparisNo !='' and YEAR(f.Tarih)=? and f.YuklemeEvrakID=50 and f.SiparisFaturaTurID=13
                                                group by k.FirmaID
                                                order by TotalNavlun desc
                                            """,year)
            liste = list()
            for item in result:
                model = NavlunMasraflarModel()
                model.firma = item.firma
                model.totalNavlun = item.TotalNavlun
                liste.append(model)
            schema = NavlunMasraflarSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getNavlunMasraflar Hata',str(e))
            
    def getNavlunMasraflarExcel(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/navlunMasraflar.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/navlunMasraflar.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')    
            satir = 2

            for item in data:

                sayfa.cell(satir,column=1,value=item['firma'])
                sayfa.cell(satir,column=2,value=item['totalNavlun'])
                satir += 1
                
        
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('getNavlunMasraflarExcel  Hata : ',str(e))
            return False
            
class NavlunMasraflarSchema(Schema):
    firma = fields.String()
    totalNavlun = fields.Float()
    
class NavlunMasraflarModel:
    firma = ""
    totalNavlun = 0
    