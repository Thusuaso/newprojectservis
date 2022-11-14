from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil



class ExcelCiktiIslem:

    def uretim_rapor_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/uretim_listesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/uretim_listesi.xlsx'

            shutil.copy2(source_path, target_path)
            

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')

            satir = 3

            for item in data_list:
               
                sayfa.cell(satir,column=1,value=item['tarih'])  
                sayfa.cell(satir,column=2,value=item['kimden'])             
                sayfa.cell(satir,column=3,value=item['kategori'])
                sayfa.cell(satir,column=4,value=item['kasano'])
                sayfa.cell(satir,column=5,value=item['urunadi'])
                sayfa.cell(satir,column=6,value=item['ocakadi'])
                sayfa.cell(satir,column=7,value=item['yuzeyadi'])
                sayfa.cell(satir,column=8,value=item['en'])
                sayfa.cell(satir,column=9,value=item['boy'])
                sayfa.cell(satir,column=10,value=item['kenar'])
                sayfa.cell(satir,column=11,value=item['adet'])
                sayfa.cell(satir,column=12,value=item['miktar'])
                sayfa.cell(satir,column=13,value=item['birimadi'])
                sayfa.cell(satir,column=14,value=item['siparisno'])

                satir += 1
            
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False

    def sevkiyat_rapor_ciktisi(self,data_list):
        try:
            source_path = 'resource_api/raporlar/sablonlar/sevkiyat_listesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/sevkiyat_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)

            sayfa = kitap.get_sheet_by_name('Sheet')

            satir = 2
            
            for item in data_list:
                
                sayfa.cell(satir,column=1,value=item['tarih'])
                sayfa.cell(satir,column=2,value=item['musteriadi'])   
                sayfa.cell(satir,column=3,value=item['kimden'])             
                sayfa.cell(satir,column=4,value=item['kategori'])
                sayfa.cell(satir,column=5,value=item['kasano'])
                sayfa.cell(satir,column=6,value=item['urunadi'])
                sayfa.cell(satir,column=7,value=item['ocakadi'])
                sayfa.cell(satir,column=8,value=item['yuzeyadi'])
                sayfa.cell(satir,column=9,value=item['en'])
                sayfa.cell(satir,column=10,value=item['boy']) 
                sayfa.cell(satir,column=11,value=item['kenar'])
                sayfa.cell(satir,column=12,value=item['adet'])
                sayfa.cell(satir,column=13,value=item['kutuadet'])
                sayfa.cell(satir,column=14,value=item['miktar'])
                sayfa.cell(satir,column=15,value=item['birimadi'])
                sayfa.cell(satir,column=16,value=item['siparisno'])
                sayfa.cell(satir,column=17,value=item['birimfiyat'])
                sayfa.cell(satir,column=18,value=item['toplam'])
                    
                satir += 1
                
                
            kitap.save(target_path)
            kitap.close()
            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False

    def seleksiyon_rapor_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/seleksiyon_listesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/seleksiyon_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=1,value=item['kasa_no'])
                sayfa.cell(satir,column=2,value=item['tarih'])
                sayfa.cell(satir,column=3,value=item['kategori'])
                sayfa.cell(satir,column=4,value=item['ocak'])
                sayfa.cell(satir,column=5,value=item['tedarikci'])
                sayfa.cell(satir,column=6,value=item['urunadi'])
                sayfa.cell(satir,column=7,value=item['kenarislem'])
                sayfa.cell(satir,column=8,value=item['en'])
                sayfa.cell(satir,column=9,value=item['boy'])
                sayfa.cell(satir,column=10,value=item['kenar'])
                sayfa.cell(satir,column=11,value=item['kutuadet'])
                sayfa.cell(satir,column=12,value=item['m2'])
                sayfa.cell(satir,column=13,value=item['adet'])
                sayfa.cell(satir,column=14,value=item['mt'])
                sayfa.cell(satir,column=15,value=item['ton'])
                sayfa.cell(satir,column=16,value=item['siparisaciklama'])
                sayfa.cell(satir,column=17,value=item['aciklama'])

               
                

                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False


    def seleksiyon_etiket_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/Seleksiyon Üretim Etiket.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/Seleksiyon Üretim Etiket.xlsx'

            shutil.copy2(source_path, target_path)
            hizala_orta = Alignment(horizontal="center")
            kalin = Font(bold=True,size=16)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sheet')
            cift_kenarlik = Side(border_style="double")
            kenarlik = Border(top=cift_kenarlik, 
                        right=cift_kenarlik,
                        bottom=cift_kenarlik,
                       left=cift_kenarlik)            
            satir = 2

            for item in data_list:

                sayfa.cell(satir,column=2,value=item['kasano']).border= kenarlik
                sayfa.cell(satir,column=2,value=item['kasano']).font = kalin
                sayfa.cell(satir,column=2,value=item['kasano']).alignment  = hizala_orta
                sayfa.cell(satir,column=3,value=item['siparisno']).border = kenarlik
                sayfa.cell(satir,column=3,value=item['siparisno']).font = kalin
                sayfa.cell(satir,column=3,value=item['siparisno']).alignment  = hizala_orta
                satir += 1
                sayfa.cell(satir ,column=2,value="Ürün Adı").border = kenarlik
                sayfa.cell(satir ,column=2,value="Ürün Adı").font = kalin
                sayfa.cell(satir ,column=2,value="Ürün Adı").alignment  = hizala_orta
                sayfa.cell(satir,column=3,value=item['urunadi']).border = kenarlik
                sayfa.cell(satir,column=3,value=item['urunadi']).font = kalin
                sayfa.cell(satir,column=3,value=item['urunadi']).alignment  = hizala_orta
                satir += 1
                sayfa.cell(satir,column=2, value="İşlem").border = kenarlik
                sayfa.cell(satir,column=2, value="İşlem").font = kalin
                sayfa.cell(satir,column=2, value="İşlem").alignment  = hizala_orta
                sayfa.cell(satir,column=3,value=item['yuzeyadi']).border = kenarlik
                sayfa.cell(satir,column=3,value=item['yuzeyadi']).font = kalin
                sayfa.cell(satir,column=3,value=item['yuzeyadi']).alignment  = hizala_orta
                satir += 1
                sayfa.cell(satir,column=2,value="Ebat").border = kenarlik
                sayfa.cell(satir,column=2,value="Ebat").font = kalin
                sayfa.cell(satir,column=2,value="Ebat").alignment  = hizala_orta
                sayfa.cell(satir,column=3,value=item['en'] +"x"+ item['boy'] +"x"+ item['kenar']).border = kenarlik
                sayfa.cell(satir,column=3,value=item['en'] +"x"+ item['boy'] +"x"+ item['kenar']).font = kalin
                sayfa.cell(satir,column=3,value=item['en'] +"x"+ item['boy'] +"x"+ item['kenar']).alignment  = hizala_orta
                satir += 1
                sayfa.cell(satir,column=2, value="Üretim Tarihi").border = kenarlik
                sayfa.cell(satir,column=2, value="Üretim Tarihi").font = kalin
                sayfa.cell(satir,column=2, value="Üretim Tarihi").alignment  = hizala_orta
                sayfa.cell(satir,column=3,value=item['tarih']).border = kenarlik
                sayfa.cell(satir,column=3,value=item['tarih']).font = kalin
                sayfa.cell(satir,column=3,value=item['tarih']).alignment  = hizala_orta
                satir += 1
                sayfa.cell(satir,column=2, value="Tedarikçi").border = kenarlik
                sayfa.cell(satir,column=2, value="Tedarikçi").font = kalin
                sayfa.cell(satir,column=2, value="Tedarikçi").alignment  = hizala_orta
                sayfa.cell(satir,column=3,value=item['kimden']).border = kenarlik
                sayfa.cell(satir,column=3,value=item['kimden']).font = kalin
                sayfa.cell(satir,column=3,value=item['kimden']).alignment  = hizala_orta
               
               
                

               
                

                satir += 2

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False


    
    
    def stok_rapor_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/Stok_listesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/Stok_listesi.xlsx'

            shutil.copy2(source_path, target_path)
            rgb=[204,102,0]
            color_string="".join([str(hex(i))[2:].upper().rjust(2, "0") for i in rgb])
           

            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            kalin = Font(bold=True,size=16)
            satir = 1
            sayfa.cell(satir,column=1,value='Ebat').fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=1,value='Ebat').font = kalin
            sayfa.cell(satir,column=2,value= "Yuzey İşlem" ).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=2,value= "Yuzey İşlem" ).font = kalin
            sayfa.cell(satir,column=3,value= "Ürün Adı" ).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=3,value= "Ürün Adı" ).font = kalin
            sayfa.cell(satir,column=4,value= "Kasa" ).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=4,value= "Kasa" ).font = kalin
            sayfa.cell(satir,column=5,value= "M2" ).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
            sayfa.cell(satir,column=5,value= "M2" ).font = kalin
            satir += 1
            for item1 in data_list:
                sayfa.cell(satir,column=1,value=item1['ebat'])
                sayfa.cell(satir,column=2,value=item1['yuzeyIslem'])
                sayfa.cell(satir,column=3,value=item1['urunAdi'])
                sayfa.cell(satir,column=4,value=item1['kasaSayisi'])
                sayfa.cell(satir,column=5,value=item1['miktar'])
                satir += 1
            

            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False          

    def stok_rapor_ciktisi_ayrinti(self,data_list):
       
        try:
            source_path = 'resource_api/raporlar/sablonlar/Stok_listesi_ayrinti.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/Stok_listesi_ayrinti.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            
            satir = 2
            item = 0
            for item in data_list:
                sayfa.cell(satir,column=2,value=item['kasanotop'])
                sayfa.cell(satir,column=3,value=item['kategoritop'])
                sayfa.cell(satir,column=4,value=item['urunaditop'])
                sayfa.cell(satir,column=5,value=item['yuzeyislemtop'])
                sayfa.cell(satir,column=6,value=item['entop'])
                sayfa.cell(satir,column=7,value=item['boytop'])
                sayfa.cell(satir,column=8,value=item['kenartop'])
                sayfa.cell(satir,column=9,value=item['miktartop'])
                sayfa.cell(satir,column=10,value=item['adettop'])
                sayfa.cell(satir,column=11,value=item['birimaditop'])
                sayfa.cell(satir,column=12,value=item['kutuadettop'])
                satir += 1

            kitap.save(target_path)
            kitap.close()

            return True
        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False          
          


    def yukleme_po_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/yukleme_po_bazında.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/yukleme_po_bazında.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
            toplam = 0
            toplam1 = 0
            listeuzunluk = len(data_list)-1 
            kalin = Font(bold=True,size=12)
            for item in data_list:

               
                sayfa.cell(satir,column=1,value=item['siparis_no'])             
                sayfa.cell(satir,column=2,value=item['musteri_adii'])
                sayfa.cell(satir,column=3,value=item['marketing'])
                sayfa.cell(satir,column=4,value=item['fob'])
                sayfa.cell(satir,column=5,value=item['dtp'])
                toplam += item['fob']
                toplam1 += item['dtp']

                satir += 1
              
            sayfa.cell(listeuzunluk+4,column=4,value=toplam).font = kalin
            sayfa.cell(listeuzunluk+4,column=5,value=toplam1).font = kalin  
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False         

    def yukleme_musteri_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/Yukleme-Musteri.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/Yukleme-Musteri.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 3
            toplam = 0
            toplam1 = 0
            listeuzunluk = len(data_list)-1 
            kalin = Font(bold=True,size=12)
            for item in data_list:

               
                         
                sayfa.cell(satir,column=1,value=item['musteri_adii'])
                sayfa.cell(satir,column=2,value=item['marketing'])
                sayfa.cell(satir,column=3,value=item['fob'])
                sayfa.cell(satir,column=4,value=item['dtp'])
                toplam += item['fob']
                toplam1 += item['dtp']

                satir += 1
              
            sayfa.cell(listeuzunluk+4,column=3,value=toplam).font = kalin
            sayfa.cell(listeuzunluk+4,column=4,value=toplam1).font = kalin  
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False         
        
    def yukleme_Yil_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/Yukleme-Yil.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/Yukleme-Yil.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
            toplam = 0
            toplam1 = 0
            listeuzunluk = len(data_list)-1 
            kalin = Font(bold=True,size=12)
            for item in data_list:

               
                         
                sayfa.cell(satir,column=1,value=item['musteri_adi'])
                sayfa.cell(satir,column=2,value=item['marketing'])
                sayfa.cell(satir,column=3,value=item['fob'])
                sayfa.cell(satir,column=4,value=item['dtp'])
                toplam += item['fob']
                toplam1 += item['dtp']

                satir += 1
              
            sayfa.cell(listeuzunluk+4,column=3,value=toplam).font = kalin
            sayfa.cell(listeuzunluk+4,column=4,value=toplam1).font = kalin  
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False         
        
    def siparis_ozet_rapor_ciktisi(self,data_list):

        try:
            source_path = 'resource_api/raporlar/sablonlar/SiparisOzet.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/SiparisOzet.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')
            satistop =0
            satir = 1
            sayfa.cell(satir,column=1,value= "Tarih")   
            sayfa.cell(satir,column=2,value="Müşteri")             
            sayfa.cell(satir,column=3,value="Po")
            sayfa.cell(satir,column=4,value="Teslim")
            sayfa.cell(satir,column=5,value="Satış Toplamı")
            sayfa.cell(satir,column=6,value="Navlun")
            sayfa.cell(satir,column=7,value="Detay 1")
            sayfa.cell(satir,column=8,value="Detay 2")
            sayfa.cell(satir,column=9,value="Detay 3")
            sayfa.cell(satir,column=10,value="Detay 4")
            sayfa.cell(satir,column=11,value="Toplam")
            satir += 1
            for item in data_list:

                sayfa.cell(satir,column=1,value=item['tarih'])   
                sayfa.cell(satir,column=2,value=item['musteri'])             
                sayfa.cell(satir,column=3,value=item['siparisnumarasi'])
                sayfa.cell(satir,column=4,value=item['teslim'])
                sayfa.cell(satir,column=5,value=item['satistoplam'])
                sayfa.cell(satir,column=6,value=item['navlun'])
                sayfa.cell(satir,column=7,value=item['detay1'])
                sayfa.cell(satir,column=8,value=item['detay2'])
                sayfa.cell(satir,column=9,value=item['detay3'])
                sayfa.cell(satir,column=10,value=item['detay4'])
                sayfa.cell(satir,column=11,value=item['detay4']+item['detay3']+item['detay2']+item['detay1']+item['navlun']+item['satistoplam'])
                satistop += item['satistoplam']

                satir += 1
            sayfa.cell(satir+1,column=5,value=satistop)
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False
      
    def depo_stok_ciktisi(self,data_list):
     
        try:
            source_path = 'resource_api/raporlar/sablonlar/Atlanta_SM_stock.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/Atlanta_SM_stock.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 3

            for item in data_list:
               
                sayfa.cell(satir,column=1,value=item['sku'])
                sayfa.cell(satir,column=2,value=item['kod'])
                sayfa.cell(satir,column=3,value=item['po']) 
                sayfa.cell(satir,column=4,value=item['tanim'])  
                 
                sayfa.cell(satir,column=5,value=item['kasa_adet'])             
                sayfa.cell(satir,column=6,value=item['kasa_kutu'])
                sayfa.cell(satir,column=7,value=item['kasa_m2'])
                sayfa.cell(satir,column=8,value=item['kasa_Sqft'])
                sayfa.cell(satir,column=9,value=item['kutu_adet'])
                sayfa.cell(satir,column=10,value=item['stok_kutu'])
                sayfa.cell(satir,column=11,value=item['stok_m2'])
                sayfa.cell(satir,column=12,value=item['stok_sqft']) 
                sayfa.cell(satir,column=13,value=item['su_kutu'])
                sayfa.cell(satir,column=14,value=item['su_m2'])
                sayfa.cell(satir,column=15,value=item['su_sqft'])
                sayfa.cell(satir,column=16,value=item['toplam_mekus'])
                sayfa.cell(satir,column=17,value=item['mekmar_fiyat'])
                sayfa.cell(satir,column=18,value=item['maya_fiyat'])
                sayfa.cell(satir,column=19,value=item['villo_fiyat'])
                sayfa.cell(satir,column=20,value=item['bd_fiyat'])

              
   
       

                satir += 1
           
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem depoCikti Hata : ',str(e))
            return False     

    
    def urunler_uretim_excel(self,data_list):
     
        try:
            source_path = 'resource_api/raporlar/sablonlar/uretilen_listesi.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/uretilen_listesi.xlsx'

            shutil.copy2(source_path, target_path)


            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')

            satir = 2
            for item in data_list:
               
                sayfa.cell(satir,column=1,value=item['kategori'])
                sayfa.cell(satir,column=2,value=item['urunAdi'])
                sayfa.cell(satir,column=3,value=item['yuzey']) 
                sayfa.cell(satir,column=4,value=item['en']) 
                sayfa.cell(satir,column=5,value=item['boy'])  
                sayfa.cell(satir,column=6,value=item['kenar'])  
                sayfa.cell(satir,column=7,value=item['sipMiktari'])  
                sayfa.cell(satir,column=8,value=item['uretimMiktar'])
                sayfa.cell(satir,column=9,value=item['uretilmesiGereken'])
                  
                

                satir += 1
           
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('ExcelCiktiIslem urunler_uretim_excel Hata : ',str(e))
            return False     
