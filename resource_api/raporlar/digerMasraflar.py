from helpers import SqlConnect,TarihIslemler
from models.raporlar import *
from marshmallow import Schema,fields
from openpyxl import *
from openpyxl.styles import Border, Side , Font ,Alignment,PatternFill
from openpyxl.cell import Cell
import shutil

class DigerMasraflar():
    def __init__(self):
        self.data = SqlConnect().data
    def getDigerMasraflar(self,year):
        try:
            satisToplami = self.data.getStoreList("""
                                                select 
                                                    s.MusteriID,
                                                    sum(su.SatisToplam) as SatisToplami,
                                                    (select m.FirmaAdi from MusterilerTB m where m.ID = s.MusteriID) as Musteri
                                                from
                                                    SiparislerTB s
                                                    inner join SiparisUrunTB su on su.SiparisNo = s.SiparisNo

                                                    where YEAR(s.YuklemeTarihi) = ?

                                                group by 
                                                    s.MusteriID
                                            """,year)
            
            
            digerMasraflar = self.data.getStoreList("""
                                                        select 
                                                            s.MusteriID,
                                                            (select m.FirmaAdi from MusterilerTB m where m.ID = s.MusteriID) as Musteri,
                                                            sum(s.NavlunSatis) as NavlunSatis,
                                                            sum(s.DetayTutar_1) as DetayTutar_1,
                                                            sum(s.DetayTutar_2) as DetayTutar_2,
                                                            sum(s.DetayTutar_3) as DetayTutar_3
                                                        from
                                                            SiparislerTB s

                                                            where  s.SiparisDurumID=3 and YEAR(s.YuklemeTarihi) = ? 

                                                        group by 
                                                            s.MusteriID
                                                    
                                                    
                                                    """,year)
            
            
            liste = list()
            for item in zip(satisToplami,digerMasraflar):
                model = DigerMasraflarModel()
                model.musteriAdi = item[0].Musteri
                model.siparisTotal = item[0].SatisToplami
                model.navlunSatis = item[1].NavlunSatis
                model.detay_1 = item[1].DetayTutar_1
                model.detay_2 = item[1].DetayTutar_2
                model.detay_3 = item[1].DetayTutar_3
                liste.append(model)
                
            schema = DigerMasraflarSchema(many=True)
            return schema.dump(liste)
                
        except Exception as e:
            print('getDigerMasraflar Hata',str(e))
            
    def getDigerMasraflarExcel(self,data):
        try:
            source_path = 'resource_api/raporlar/sablonlar/digerMasraflar.xlsx'
            target_path = 'resource_api/raporlar/dosyalar/digerMasraflar.xlsx'

            shutil.copy2(source_path, target_path)
            kitap = load_workbook(target_path)
            sayfa = kitap.get_sheet_by_name('Sayfa1')    
            satir = 2

            for item in data:

                sayfa.cell(satir,column=1,value=item['musteriAdi'])
                sayfa.cell(satir,column=2,value=item['siparisTotal'])
                sayfa.cell(satir,column=3,value=item['navlunSatis'])
                sayfa.cell(satir,column=4,value=item['detay_1'])
                sayfa.cell(satir,column=5,value=item['detay_2'])
                sayfa.cell(satir,column=6,value=item['detay_3'])
                satir += 1
                
        
            kitap.save(target_path)
            kitap.close()

            return True

        except Exception as e:
            print('getDigerMasraflarExcel  Hata : ',str(e))
            return False
            
class DigerMasraflarSchema(Schema):
    musteriAdi = fields.String()
    siparisTotal = fields.Float()
    navlunSatis = fields.Float()
    detay_1 = fields.Float()
    detay_2 = fields.Float()
    detay_3 = fields.Float()
    
class DigerMasraflarModel:
    musteriAdi = ""
    siparisTotal = 0
    navlunSatis = 0
    detay_1 = 0
    detay_2 = 0
    detay_3 = 0
    